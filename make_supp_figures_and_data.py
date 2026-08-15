"""
make_supp_figures_and_data.py
  -- supplementary figures (Fig. S1-S4) + raw-data workbooks for EVERY figure

Two jobs:
 1. Plot the data behind Supplementary Tables S3/S4/S6/S7 as figures S1-S4.
 2. Export the exact arrays drawn in every figure -- main Fig. 1-5 and
    supplementary Fig. S1-S4 -- as one Excel workbook per figure, named after
    the figure (fig2_achievable_region.xlsx belongs to
    fig2_achievable_region.png), one sheet per panel, so the figures can be
    re-plotted or restyled in Excel/Origin without touching the .mat archives.

It imports make_figures, so every exported array is the same object the PNG was
drawn from -- the two cannot drift apart. Run AFTER make_figures.py:

    python3 make_figures.py
    python3 make_supp_figures_and_data.py
"""
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from scipy.io import loadmat
from openpyxl import Workbook

import make_figures as MF   # loads all archives, defines helpers & plotted arrays

# 추가 아카이브 (보충 그림 전용)
CAL = loadmat('calibrate_random_cost.mat')
CVG = loadmat('convergence_check_result.mat')
CFM = loadmat('reeval_confirm_2040_result.mat')
P25 = loadmat('patch_convergence_result.mat')     # 15/25/35 mm
P100 = loadmat('patch_convergence_100.mat')       # 100 mm follow-up

BANDS = MF.BANDS
S_LAMB = MF.S_LAMB


# ============================================================
#  Excel writer: one workbook per figure, one sheet per panel
# ============================================================
def wb_write(stem, sheets):
    """sheets: {sheet_name: (headers, columns)} -- columns may differ in length."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, (headers, cols) in sheets.items():
        ws = wb.create_sheet(name[:31])
        ws.append(list(headers))
        cols = [np.asarray(c, dtype=object).ravel() for c in cols]
        n = max(len(c) for c in cols)
        for i in range(n):
            row = []
            for c in cols:
                if i < len(c):
                    v = c[i]
                    if isinstance(v, (np.floating, np.integer)):
                        v = v.item()
                    if isinstance(v, float) and not np.isfinite(v):
                        v = None
                    row.append(v)
                else:
                    row.append(None)
            ws.append(row)
    wb.save(stem + '.xlsx')
    print(f'  data  -> {stem}.xlsx  ({len(sheets)} sheets)')


# ============================================================
#  Raw data for the MAIN figures (arrays identical to make_figures.py)
# ============================================================
def export_fig1():
    t = np.linspace(0, 1, 200)
    z_hemi = np.sqrt(1 - t**2)
    tt, z_ff = MF.profile(MF.band_x[2, :])
    cx = np.sort(MF.band_x[2, 0:5]); cy = MF.band_x[2, 5:10] * MF.band_x[2, 12]
    wb_write('fig1_platform', {
        'b_lens_profiles': (['r_over_rlens', 'z_hemisphere_unit', 'r_freeform', 'z_freeform_4060opt'],
                            [t, z_hemi, tt, z_ff]),
        'b_control_points': (['x_ctrl', 'z_ctrl (40-60 optimum)'], [cx, cy]),
        'c_band_definition': (['band', 'theta_lo_deg', 'theta_hi_deg', 'S_Lambertian'],
                              [BANDS, [0, 20, 40, 60], [20, 40, 60, 80], S_LAMB]),
        'readme': (['note'], [[
            'Panels (a) and (d) are schematics with no plotted data.',
            'b: profiles in units of the lens radius; freeform = 40-60 deg optimum.',
            'Source archives: opt_4band_result_25by25.mat']]),
    })


def export_fig2():
    S_win = MF.band_eqe / MF.band_tot
    E_max_c = float(np.max(MF.Tc))
    gain_sel = S_win / MF.S_nat
    gain_tot = MF.band_tot / E_max_c
    gain_net = MF.band_eqe / (MF.S_nat * E_max_c)

    pf = np.polyfit(MF.Tp, MF.Bp[:, 2], 1)
    r2 = np.corrcoef(MF.Tp, MF.Bp[:, 2])[0, 1] ** 2

    th = np.linspace(0, 1, 300)
    hx = np.concatenate(([0.0], MF.HEMI_X, [1.0]))
    hy = np.concatenate(([1.0], MF.HEMI_Y, [0.0]))
    from scipy.interpolate import PchipInterpolator
    k_ht = int(np.argmax(MF.hemi_tot))
    z_h = PchipInterpolator(hx, hy)(th) * MF.hemi_x[k_ht, 2]
    profs = {f'z_{BANDS[j]}': MF.profile(MF.band_x[j, :])[1] for j in range(4)}

    ws_val = MF.W['ws_val'].ravel()
    gbest = float(np.asarray(MF.F['gBestEQE']).ravel()[0])
    ff = np.fmax(np.concatenate((MF.band_eqe, [gbest])), np.nan_to_num(ws_val, nan=-np.inf))
    hm = MF.hemi_val[:5]
    base = MF.W['base_val'].ravel(); sdb = MF.W['base_sd'].ravel(); sdw = MF.W['ws_sd'].ravel()
    se = np.sqrt(sdb**2 + sdw**2) / np.sqrt(3)
    tval = (ws_val - base) / se
    rel = 100 * (ws_val - base) / base

    wb_write('fig2_achievable_region', {
        'a_selectivity': (['band', 'S_dedicated', 'S_natural_median', 'S_natural_p10',
                           'S_natural_p90', 'S_Lambertian'],
                          [BANDS, S_win, MF.S_nat, MF.S_lo, MF.S_hi, S_LAMB]),
        'b_gain_decomposition': (['band', 'selectivity_gain', 'total_EQE_ratio', 'net_band_gain'],
                                 [BANDS, gain_sel, gain_tot, gain_net]),
        'c_collapse_points': (['EQE_total', 'EQE_40_60', 'phase (1=random)'],
                              [MF.Tp, MF.Bp[:, 2], MF.phase_p]),
        'c_fit_and_optima': (['fit_slope', 'fit_intercept', 'fit_R2',
                              'pareto_EQE_total', 'pareto_EQE_40_60'],
                             [[pf[0]], [pf[1]], [r2],
                              MF.P['pareto_tot'].ravel(), MF.P['pareto_band'].ravel()]),
        'd_profiles': (['r_over_rlens', 'z_hemisphere_opt'] + list(profs.keys()),
                       [th, z_h] + list(profs.values())),
        'e_Gj': (['objective', 'freeform_best', 'hemisphere', 'G_j'],
                 [BANDS + ['total'], ff, hm, ff / hm]),
        'f_warmstart_control': (['arm', 'hemisphere_val', 'hemisphere_sd', 'warmstart_val',
                                 'warmstart_sd', 'gain_pct', 't_value', 't_threshold'],
                                [BANDS + ['total'], base, sdb, ws_val, sdw, rel, tval,
                                 [2.132]*5]),
    })


def export_fig3():
    sheets = {}
    for j in range(4):
        S = MF.Bp[:, j] / MF.Tp
        pf = np.polyfit(MF.Tp, S, 1)
        sheets[f'band_{BANDS[j]}'] = (
            ['EQE_total', f'S_{BANDS[j]}', 'phase (1=random)'],
            [MF.Tp, S, MF.phase_p])
        sheets[f'band_{BANDS[j]}_meta'] = (
            ['Pearson_R', 'fit_slope', 'fit_intercept', 'S_Lambertian'],
            [[MF.R_full[j]], [pf[0]], [pf[1]], [S_LAMB[j]]])
    wb_write('fig3_selectivity_map', sheets)


def export_fig4():
    d1, d3 = MF.d1, MF.d3
    a_list = np.array([0.0, 0.01, 0.02, 0.05, 0.10, 0.20]) * 100
    ideal_b = np.array([100.0, 94.2, 89.1, 76.6, 62.1, 45.0])
    flat_b = np.array([33.8, 33.3, 32.8, 31.3, 29.1, 25.5])
    WALL = float(d1['wall']) * 100
    th = d1['th']
    Ti = np.zeros_like(th)
    lo, hi = float(d1['th_sub_lo']), float(d1['th_sub_hi'])
    Ti[(th >= lo) & (th <= hi)] = 1
    wb_write('fig4_recycling_routes', {
        'a_loss_dependence': (['loss_a_pct', 'ideal_filter_pct', 'planar_pct', 'single_pass_wall_pct'],
                              [a_list, ideal_b, flat_b, [WALL]*len(a_list)]),
        'b_bandwidth': (['dlam_nm', 'band_delivery_pct', 'selectivity_pct', 'single_pass_wall_pct'],
                        [d3['dlam'], d3['band']*100, d3['sel']*100, [WALL]*len(d3['dlam'])]),
        'c_angular_response': (['theta_substrate_deg', 'T_planar', 'T_DBR_8pair', 'T_ideal',
                                'window_lo_deg', 'window_hi_deg'],
                               [th, d1['T_flat'], d1['T_dbr'], Ti, [lo], [hi]]),
        'readme': (['note'], [['Panel (d) is a schematic with no plotted data.',
                               'Source: angular_recycling_result.npz, angular_recycling_bandwidth.npz']]),
    })


def export_fig5():
    fams = [(MF.C, 'convex'), (MF.I, 'inverted'), (MF.Rn, 'random')]
    S_conv, _, _ = MF.natural_composition(*MF.clean_log(MF.C['EVAL_LOG']))
    R_conv = MF.band_corr(*MF.clean_log(MF.C['EVAL_LOG']))
    sheets = {}
    for D, name in fams:
        T, B = MF.clean_log(D['EVAL_LOG'])
        Snat, _, _ = MF.natural_composition(T, B)
        Smean = np.mean(B / T[:, None], axis=0)
        Rsel = MF.band_corr(T, B)
        pf = np.polyfit(T, B[:, 2], 1)
        r2 = np.corrcoef(T, B[:, 2])[0, 1] ** 2
        sheets[f'{name}_1_collapse'] = (['EQE_total', 'EQE_40_60'], [T, B[:, 2]])
        sheets[f'{name}_1_fit'] = (['slope', 'intercept', 'R2', 'n'], [[pf[0]], [pf[1]], [r2], [T.size]])
        sheets[f'{name}_2_composition'] = (
            ['band', 'S_top20_median', 'S_population_mean', 'S_convex_ref', 'S_Lambertian'],
            [BANDS, Snat, Smean, S_conv, S_LAMB])
        sheets[f'{name}_3_drift'] = (['band', 'R_this_family', 'R_convex_ref'],
                                     [BANDS, Rsel, R_conv])
    wb_write('fig5_families', sheets)


# ============================================================
#  Supplementary figures
# ============================================================
def figS1_patch():
    """Table S7 as a figure: total EQE vs patch + selectivity vs patch."""
    Pz = list(P25['PATCHES'].ravel()) + list(P100['PATCHES'].ravel())
    Et = [P25['Et'][i] for i in range(3)] + [P100['Et'].ravel()]
    m = np.array([np.nanmean(e) for e in Et])
    sd = np.array([np.nanstd(e, ddof=1) for e in Et])
    Bn25 = P25['Bn']; Bn100 = P100['Bn']
    S = np.vstack([np.nanmean(Bn25[i], axis=1) / m[i] for i in range(3)] +
                  [np.nanmean(Bn100[0], axis=1) / m[3]])

    fig, axs = plt.subplots(1, 2, figsize=(7.2, 2.9))
    ax = axs[0]
    ax.errorbar(Pz, m, yerr=sd, fmt='o-', lw=1.6, ms=5, color='#3B6EA5', capsize=3)
    ax.axhline(0.54679, color='#C6512F', ls='--', lw=1.0,
               label='hemisphere @ 25 mm (0.5468)')
    for x, y in zip(Pz, m):
        dy = 7 if y < 0.53 else -11
        ax.annotate(f'{y:.4f}', (x, y), textcoords='offset points', xytext=(6, dy), fontsize=6.4)
    ax.set_xlabel('patch size (mm)'); ax.set_ylabel(r'EQE$_{\rm total}$')
    ax.set_title('(a) total EQE keeps rising with patch size', loc='left')
    ax.set_xlim(5, 110); ax.legend(fontsize=6.4, loc='lower right'); ax.grid(alpha=0.25)

    ax = axs[1]
    for j in range(4):
        ax.plot(Pz, S[:, j], 'o-', lw=1.4, ms=4, color=MF.CB[j], label=BANDS[j])
    ax.set_xlabel('patch size (mm)'); ax.set_ylabel(r'selectivity $S_j$')
    ax.set_title('(b) the composition does not move', loc='left')
    ax.set_xlim(5, 110); ax.set_ylim(0, 0.45)
    ax.legend(fontsize=6.2, ncol=4, loc='upper center', columnspacing=0.9)
    ax.grid(alpha=0.25)
    fig.tight_layout()
    MF.save(fig, 'figS1_patch_dependence')

    reps = np.full((4, 3), np.nan)
    for i, e in enumerate(Et):
        v = np.asarray(e).ravel(); reps[i, :len(v)] = v[:3]
    wb_write('figS1_patch_dependence', {
        'a_total_EQE': (['patch_mm', 'rep1', 'rep2', 'rep3', 'mean', 'sd',
                         'hemisphere_ref_25mm'],
                        [Pz, reps[:, 0], reps[:, 1], reps[:, 2], m, sd, [0.54679]]),
        'b_selectivity': (['patch_mm'] + [f'S_{b}' for b in BANDS],
                          [Pz] + [S[:, j] for j in range(4)]),
    })


def figS2_warmstart():
    """Table S6 as a figure: absolute values per arm + the 5-repeat confirmation."""
    base = MF.W['base_val'].ravel(); sdb = MF.W['base_sd'].ravel()
    ws = MF.W['ws_val'].ravel(); sdw = MF.W['ws_sd'].ravel()
    se = np.sqrt(sdb**2 + sdw**2) / np.sqrt(3)
    tval = (ws - base) / se
    labels = BANDS + ['total']

    fig, axs = plt.subplots(1, 2, figsize=(7.2, 2.9))
    ax = axs[0]
    xx = np.arange(5)
    ax.bar(xx - 0.19, base, 0.38, yerr=sdb, capsize=2, color='#3B6EA5', ec='k', lw=0.4,
           label='hemisphere (re-measured)')
    ax.bar(xx + 0.19, ws, 0.38, yerr=sdw, capsize=2, color='#C6512F', ec='k', lw=0.4,
           label='restarted at hemisphere')
    for i in range(5):
        ax.text(xx[i], max(base[i], ws[i]) + 0.015,
                f'$t$={tval[i]:.0f}' if tval[i] > 10 else f'$t$={tval[i]:.1f}',
                ha='center', fontsize=6.2)
    ax.set_xticks(xx); ax.set_xticklabels(labels, fontsize=6.6)
    ax.set_ylabel('EQE at the matching objective'); ax.set_ylim(0, 0.66)
    ax.set_title('(a) hemisphere-seeded control, all arms', loc='left')
    ax.legend(fontsize=6.2, loc='upper left'); ax.grid(alpha=0.25, axis='y')

    ax = axs[1]
    vH = CFM['vH'].ravel(); vW = CFM['vW'].ravel()
    ax.plot(np.full(vH.size, 0), vH, 'o', ms=6, color='#3B6EA5', alpha=0.75)
    ax.plot(np.full(vW.size, 1), vW, 'o', ms=6, color='#C6512F', alpha=0.75)
    for m_, x in [(vH.mean(), 0), (vW.mean(), 1)]:
        ax.plot([x-0.18, x+0.18], [m_, m_], 'k-', lw=1.4)
    ax.set_xticks([0, 1]); ax.set_xticklabels(['hemisphere', 'warm start'], fontsize=7)
    ax.set_xlim(-0.6, 1.6)
    ax.set_ylabel(r'EQE$_{20-40^\circ}$')
    tc = float(CFM['tval'].ravel()[0])
    ax.set_title(f'(b) 20-40$^\\circ$ five-repeat confirmation:  $t$ = {tc:.1f}, +0.29%',
                 loc='left')
    ax.grid(alpha=0.25, axis='y')
    fig.tight_layout()
    MF.save(fig, 'figS2_warmstart_control')

    wb_write('figS2_warmstart_control', {
        'a_all_arms': (['arm', 'hemisphere_val', 'hemisphere_sd', 'warmstart_val',
                        'warmstart_sd', 'gain_pct', 't_value'],
                       [labels, base, sdb, ws, sdw, 100*(ws-base)/base, tval]),
        'b_confirmation_reps': (['rep', 'hemisphere_EQE_20_40', 'warmstart_EQE_20_40'],
                                [np.arange(1, vH.size+1), vH, vW]),
        'b_confirmation_stats': (['mean_hemi', 'sd_hemi', 'mean_warm', 'sd_warm',
                                  't_value', 't_threshold_df8'],
                                 [[vH.mean()], [vH.std(ddof=1)], [vW.mean()],
                                  [vW.std(ddof=1)], [tc], [1.860]]),
    })


def figS3_calibration():
    """Table S3 as a figure: cost vs accuracy of the supercell reductions."""
    names = ['reference', 'half rays', 'coarser wavelength', 'coarser grid',
             'fewer lenslets', 'adopted combination']
    t = CAL['tSec'].ravel(); E = CAL['Etot'].ravel(); S = CAL['Sel']
    dS = np.max(np.abs(S - S[0]), axis=1) * 100

    fig, ax = plt.subplots(figsize=(6.4, 3.0))
    xx = np.arange(len(names))
    b = ax.bar(xx, t, 0.55, color=['0.65']*5 + ['#4F8F5B'], ec='k', lw=0.5)
    ax.set_ylabel('evaluation time (s)')
    ax.set_xticks(xx)
    ax.set_xticklabels(names, fontsize=6.4, rotation=18, ha='right')
    for i in range(len(names)):
        ax.text(xx[i], t[i] + 12, f'{t[0]/t[i]:.2f}x', ha='center', fontsize=6.4)
    ax2 = ax.twinx()
    ax2.plot(xx[1:], dS[1:], 'o--', color='#C6512F', ms=5, lw=1.2)
    ax2.axhline(0.5, color='#C6512F', ls=':', lw=1.0)
    ax2.set_ylabel('max |Δ selectivity| (pp)', color='#C6512F')
    ax2.tick_params(axis='y', labelcolor='#C6512F'); ax2.set_ylim(0, 1.0)
    ax.set_title('supercell cost calibration: 4.3x faster at 0.39 pp deviation', loc='left')
    fig.tight_layout()
    MF.save(fig, 'figS3_cost_calibration')

    wb_write('figS3_cost_calibration', {
        'calibration': (['setting', 'lenslets', 'grid', 'rays', 'lambda_step_nm',
                         'time_s', 'speedup', 'EQE_total', 'max_dSel_pp'],
                        [names, [8, 8, 8, 8, 6, 6], [201, 201, 201, 141, 201, 141],
                         [10000, 5000, 10000, 10000, 10000, 5000],
                         [10, 10, 20, 10, 10, 20], t, t[0]/t, E, dS]),
        'selectivity_by_setting': (['setting'] + [f'S_{b}' for b in BANDS],
                                   [names] + [S[:, j] for j in range(4)]),
    })


def figS4_convergence():
    """Table S4 as a figure: drift correlations under stricter evaluation."""
    R0 = CVG['R_lo'].ravel(); R1 = CVG['R_A_sm'].ravel(); R2 = CVG['R_B_sm'].ravel()
    fig, ax = plt.subplots(figsize=(5.4, 3.0))
    xx = np.arange(4); w = 0.26
    ax.bar(xx - w, R0, w, color='0.70', ec='k', lw=0.4, label='baseline rays')
    ax.bar(xx, R1, w, color='#3B6EA5', ec='k', lw=0.4, label=r'20$\times$ rays, $\times$3')
    ax.bar(xx + w, R2, w, color='#4F8F5B', ec='k', lw=0.4, label='broadband 450-750 nm')
    ax.axhline(0, color='k', lw=0.8)
    ax.set_xticks(xx); ax.set_xticklabels(BANDS)
    ax.set_ylabel(r'$R\,($EQE$_{\rm total},\,S_j)$'); ax.set_ylim(-1.0, 1.0)
    ax.set_title('drift correlations are stable under stricter evaluation', loc='left')
    ax.legend(fontsize=6.4, loc='lower left'); ax.grid(alpha=0.25, axis='y')
    fig.tight_layout()
    MF.save(fig, 'figS4_convergence')

    wb_write('figS4_convergence', {
        'R_by_condition': (['band', 'R_baseline', 'R_20x_rays', 'R_broadband'],
                           [BANDS, R0, R1, R2]),
    })


if __name__ == '__main__':
    print('\n[main-figure data]')
    export_fig1(); export_fig2(); export_fig3(); export_fig4(); export_fig5()
    print('\n[supplementary figures + data]')
    figS1_patch(); figS2_warmstart(); figS3_calibration(); figS4_convergence()
    print('\nall workbooks and supplementary figures written.')
