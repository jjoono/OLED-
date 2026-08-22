"""Publish the 2026-08-20 T/R campaign as one workbook.

Sixteen samples -- two seeds x (bare + Ag 4,5,6,7,8,10,12 nm) -- on one
wavelength grid, 350-850 nm.  Four data sheets share an identical layout so a
column can be read straight across: T, R as the instrument reported it, R after
the back-surface correction, and A.  A is written as a formula, =100-T-Rcorr,
written as computed numbers, not formulas: these are measured constants rather
than model inputs, and a value in every cell reads back correctly from pandas,
LibreOffice and Excel alike.

Three samples are incomplete and are kept as labelled columns rather than
dropped: HATCN5 bare was never run, MoOx5 bare has R but no T, MoOx5/Ag12 has
T but no R.  The back-surface correction needs T, so a missing T also empties
that sample's Rcorr and A.
"""
import csv, os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
RAW  = os.path.join(BASE, "data", "TR_20260820", "raw")
OUT  = os.path.join(BASE, "data", "TR_20260820", "TRA_16samples.xlsx")
KEEP = 0.843

SAMPLES = [("1-1","HATCN",0),  ("1-2","HATCN",4), ("1-3","HATCN",5), ("1-4","HATCN",6),
           ("2-1","HATCN",7),  ("2-2","HATCN",8), ("2-3","HATCN",10),("2-4","HATCN",12),
           ("1-9","MoOx",0),   ("1-10","MoOx",4), ("1-11","MoOx",5), ("1-12","MoOx",6),
           ("2-9","MoOx",7),   ("2-10","MoOx",8), ("2-11","MoOx",10),("2-12","MoOx",12)]

RS = {("HATCN",4):52.2, ("HATCN",5):23.3, ("HATCN",6):18.9, ("HATCN",7):12.7,
      ("HATCN",8):9.1,  ("HATCN",10):7.5, ("HATCN",12):5.3,
      ("MoOx",4):138.0, ("MoOx",5):49.1,  ("MoOx",6):32.0,  ("MoOx",7):14.9,
      ("MoOx",8):10.8,  ("MoOx",10):9.6,  ("MoOx",12):6.6}

FONT   = "Arial"
HEAD   = PatternFill("solid", fgColor="1E2761")
SUBHD  = PatternFill("solid", fgColor="DCE3F0")
WARN   = PatternFill("solid", fgColor="FFF2CC")
THIN   = Border(bottom=Side("thin", color="B0B8CC"))


def read(p):
    out = {}
    if not os.path.exists(p):
        return out
    for row in csv.reader(open(p)):
        try:
            out[float(row[0])] = float(row[1])
        except (ValueError, IndexError):
            pass
    return out


def n_glass(l):
    return 1.5220 + 3900.0 / l**2


def label(seed, d):
    return f"{seed} 5 / Ag {d}" if d else f"{seed} 5 (bare)"


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        h = ws.cell(row=1, column=c)
        h.fill, h.font = HEAD, Font(name=FONT, bold=True, color="FFFFFF", size=10)
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        s = ws.cell(row=2, column=c)
        s.fill, s.font = SUBHD, Font(name=FONT, bold=True, size=9)
        s.alignment = Alignment(horizontal="center")
        s.border = THIN
    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 13
    for c in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.row_dimensions[1].height = 30


def data_sheet(wb, title, lam, values, note):
    """values[(sid)] -> dict lambda->number, or None for a missing sample."""
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value="wavelength").alignment = Alignment(
        horizontal="center", vertical="center")
    ws.cell(row=2, column=1, value="nm")
    for j, (sid, seed, d) in enumerate(SAMPLES):
        ws.cell(row=1, column=2 + j, value=label(seed, d))
        ws.cell(row=2, column=2 + j, value=sid)
    style_header(ws, 1 + len(SAMPLES))
    for i, l in enumerate(lam):
        r = 3 + i
        c = ws.cell(row=r, column=1, value=l)
        c.font, c.number_format = Font(name=FONT, bold=True, size=10), "0"
        for j, (sid, _, _) in enumerate(SAMPLES):
            v = values.get(sid)
            if v is None or l not in v:
                continue
            cc = ws.cell(row=r, column=2 + j, value=round(v[l], 4))
            cc.font, cc.number_format = Font(name=FONT, size=10), "0.000"
    # mark the columns that carry no data at all
    for j, (sid, _, _) in enumerate(SAMPLES):
        if values.get(sid) is None or not values[sid]:
            cell = ws.cell(row=2, column=2 + j)
            cell.fill, cell.value = WARN, sid + " (없음)"
    ws.cell(row=len(lam) + 5, column=1, value=note).font = Font(
        name=FONT, italic=True, size=9, color="555555")
    return ws


def main():
    T, R = {}, {}
    for sid, seed, d in SAMPLES:
        T[sid] = read(os.path.join(RAW, f"{sid}T.csv"))
        R[sid] = read(os.path.join(RAW, f"{sid}R.csv"))

    lams = set()
    for sid in T:
        lams |= set(T[sid]) | set(R[sid])
    lam = sorted(l for l in lams if 350 <= l <= 850)

    Rc = {}
    for sid, _, _ in SAMPLES:
        out = {}
        for l, r in R[sid].items():
            t = T[sid].get(l)
            if t is None:
                continue                      # correction needs the transmitted beam
            Rb = ((n_glass(l) - 1) / (n_glass(l) + 1))**2
            out[l] = r + 100 * (1 - KEEP) * (t / 100.0)**2 * Rb
        Rc[sid] = out

    wb = Workbook()
    wb.remove(wb.active)

    # ---- README -----------------------------------------------------------
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 96
    rows = [
        ("절대 투과도 / 반사도 / 흡수도", ""),
        ("측정일", "2026-08-20"),
        ("장비", "Agilent Cary 6000i + UMA G6875A (universal measurement accessory)"),
        ("기하", "T: 수직 입사, 절대.  R: 시료 6도 / 디텍터 12도, 절대 (직진빔 100% 기준)"),
        ("파장", "350-850 nm, 2 nm 간격"),
        ("기판", "소다라임 유리 (n = 1.5220 + 3900/lambda^2, lambda in nm)"),
        ("샘플", "유리 / seed 5 nm / Ag d,  seed = HATCN 또는 MoOx,  d = 0,4,5,6,7,8,10,12 nm"),
        ("", ""),
        ("시트 구성", ""),
        ("T", "절대 투과도 (%). 장비 출력 그대로. 보정 불필요 - 수직 입사에서 기판 다중반사 성분이 동일 방향으로 나옴"),
        ("R_measured", "절대 반사도 (%). 장비 출력 그대로"),
        ("R_corrected", "후면 반사 손실 보정 후 (%). 본 파일의 권장 R 값"),
        ("A", "흡수도 (%) = 100 - T - R_corrected.  계산된 값으로 기록됨"),
        ("Summary", "550 nm 대표값 + 면저항 + 파생량"),
        ("", ""),
        ("후면 반사 보정", ""),
        ("이유", "6도 입사에서 기판 후면 반사빔이 약 0.14 mm 옆으로 밀려 검출 개구를 일부 벗어난다"),
        ("보정식", "R_corr = R_meas + 100 * (1 - 0.843) * (T/100)^2 * R_back,  R_back = ((n-1)/(n+1))^2"),
        ("계수 0.843", "맨유리에서 결정. 이 값을 쓰면 450-700 nm 흡수도가 +0.01 +/- 0.15 %p, 즉 0이 된다"),
        ("검증", "400 nm 이하 Fe3+ 흡수와 750 nm 이상 Fe2+ 꼬리는 물리적 흡수로 그대로 남는다"),
        ("", ""),
        ("결측", ""),
        ("1-1 HATCN5 bare", "미측정 (램프 고장으로 미실시)"),
        ("1-9 MoOx5 bare", "R만 있음. T 파일 미확보 -> R_corrected 와 A 계산 불가"),
        ("2-12 MoOx5/Ag12", "T만 있음. R 파일 미확보 -> A 계산 불가"),
        ("", ""),
        ("주의", ""),
        ("A 의 정의", "A = 1 - T - R 이므로 산란(S)이 포함된다. 실제로는 A+S. 각도 분해 측정 전까지 분리 불가"),
        ("두께", "QCM 공칭값. XRR/엘립소로 독립 확인 필요"),
        ("소자 내 흡수", "여기 A 는 유리/시드/Ag/공기 구조의 값. 소자 내 one-pass 흡수와 다르다 (별도 TMM)"),
    ]
    for i, (a, b) in enumerate(rows, start=1):
        ca, cb = ws.cell(row=i, column=1, value=a), ws.cell(row=i, column=2, value=b)
        bold = (b == "" and a != "")
        ca.font = Font(name=FONT, bold=True, size=12 if i == 1 else 10,
                       color="1E2761" if bold or i == 1 else "000000")
        cb.font = Font(name=FONT, size=10)
        cb.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- data sheets ------------------------------------------------------
    data_sheet(wb, "T", lam, T,
               "절대 투과도 (%), 수직 입사. 장비 출력 그대로 - 보정 없음.")
    data_sheet(wb, "R_measured", lam, R,
               "절대 반사도 (%), 시료 6도 / 디텍터 12도. 장비 출력 그대로.")
    data_sheet(wb, "R_corrected", lam, Rc,
               "후면 반사 보정 후 반사도 (%). R_corr = R_meas + 15.7% x (T/100)^2 x R_back.")

    wsA = wb.create_sheet("A")
    wsA.cell(row=1, column=1, value="wavelength").alignment = Alignment(
        horizontal="center", vertical="center")
    wsA.cell(row=2, column=1, value="nm")
    for j, (sid, seed, d) in enumerate(SAMPLES):
        wsA.cell(row=1, column=2 + j, value=label(seed, d))
        wsA.cell(row=2, column=2 + j, value=sid)
    style_header(wsA, 1 + len(SAMPLES))
    for i, l in enumerate(lam):
        r = 3 + i
        c = wsA.cell(row=r, column=1, value=l)
        c.font, c.number_format = Font(name=FONT, bold=True, size=10), "0"
        for j, (sid, _, _) in enumerate(SAMPLES):
            if l not in T[sid] or l not in Rc[sid]:
                continue
            cc = wsA.cell(row=r, column=2 + j,
                          value=round(100 - T[sid][l] - Rc[sid][l], 4))
            cc.font, cc.number_format = Font(name=FONT, size=10), "0.000"
    for j, (sid, _, _) in enumerate(SAMPLES):
        if not (T[sid] and Rc[sid]):
            cell = wsA.cell(row=2, column=2 + j)
            cell.fill, cell.value = WARN, sid + " (없음)"
    wsA.cell(row=len(lam) + 5, column=1,
             value="A = 100 - T - R_corrected (%). 산란이 포함된 값이다 (실제로는 A+S).").font = \
        Font(name=FONT, italic=True, size=9, color="555555")

    # ---- Summary ----------------------------------------------------------
    ws = wb.create_sheet("Summary")
    hdr = ["시료", "seed", "Ag 두께 (nm)", "면저항 (ohm/sq)", "저항률 (uOhm cm)",
           "T @550 (%)", "R_corr @550 (%)", "A @550 (%)",
           "T @450 (%)", "A @450 (%)", "T @650 (%)", "A @650 (%)"]
    for c, h in enumerate(hdr, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font = HEAD, Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 9
    for c in range(3, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.row_dimensions[1].height = 32

    def at(dic, l):
        return dic.get(l)

    for i, (sid, seed, d) in enumerate(SAMPLES):
        r = 2 + i
        rs = RS.get((seed, d))
        vals = [sid, seed, d if d else None, rs,
                round(rs * d * 0.1, 2) if rs else None]
        for l in (550, 450, 650):
            t, rc = at(T[sid], l), at(Rc[sid], l)
            if l == 550:
                vals += [round(t, 3) if t is not None else None,
                         round(rc, 3) if rc is not None else None,
                         round(100 - t - rc, 3) if None not in (t, rc) else None]
            else:
                vals += [round(t, 3) if t is not None else None,
                         round(100 - t - rc, 3) if None not in (t, rc) else None]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT, size=10)
            if c >= 4:
                cell.number_format = "0.00" if c != 4 else "0.0"
        if not (T[sid] and Rc[sid]):
            for c in range(1, len(hdr) + 1):
                ws.cell(row=r, column=c).fill = WARN
    n = len(SAMPLES) + 3
    notes = ["저항률 = 면저항 x 두께 x 0.1  (1 ohm nm = 0.1 uOhm cm). Ag 벌크 1.59 uOhm cm.",
             "닫힘 두께: HATCN 5 nm, MoOx 7 nm - 수송(rho = rho0 + C/d 의 꺾임)과 광학(A 평탄부의 계단) 두 경로에서 독립적으로 같은 값.",
             "노란 행 = 결측으로 A 계산 불가.",
             "MoOx/Ag 10 nm 는 면저항과 광학 양쪽에서 이상치 - 재증착 필요."]
    for k, t in enumerate(notes):
        ws.cell(row=n + k, column=1, value=t).font = Font(
            name=FONT, italic=True, size=9, color="555555")

    wb.save(OUT)
    ok = sum(1 for sid, _, _ in SAMPLES if T[sid] and Rc[sid])
    print(f"wrote {os.path.relpath(OUT, BASE)}")
    print(f"  {len(lam)} wavelengths {lam[0]:.0f}-{lam[-1]:.0f} nm, "
          f"{len(SAMPLES)} samples, {ok} complete")


if __name__ == "__main__":
    main()
