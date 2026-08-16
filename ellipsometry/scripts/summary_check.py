import openpyxl, numpy as np
import os, pathlib
DATA_DIR = pathlib.Path(os.environ.get('ELLIPS_DATA', './data'))
OUT_DIR = pathlib.Path(os.environ.get('ELLIPS_OUT', './out'))
SCRATCH_DIR = pathlib.Path(os.environ.get('ELLIPS_SCRATCH', './out'))

fp=str(DATA_DIR / 'summary.xlsx')
wb=openpyxl.load_workbook(fp,data_only=True,read_only=True)
sheets=wb.sheetnames
def tofloat(v):
    if v is None: return np.nan
    if isinstance(v,(int,float)): return float(v)
    try: return float(str(v).strip())
    except: return np.nan
data={}
for sh in sheets:
    ws=wb[sh]
    arr=[]
    for i,row in enumerate(ws.iter_rows(min_row=3,values_only=True)):
        arr.append([tofloat(v) for v in (list(row)+[None]*50)[:50]])
    data[sh]=np.array(arr,dtype=float)
wb.close()

WLC=[0,11,22,38,44]
issues=[]
ref_wl=None
for sh in sheets:
    A=data[sh]
    wl=A[:,0]
    ok=np.isfinite(wl)
    n_ok=int(ok.sum())
    # 1) numeric coverage in Psi/Delta block
    pd_nan=int(np.sum(~np.isfinite(A[np.where(ok)[0]][:,1:11])))
    # 2) wavelength grid consistency within sheet
    wl_mis=[]
    for c in WLC[1:]:
        diff=np.nanmax(np.abs(A[:,c]-wl)) if np.isfinite(A[:,c]).any() else np.nan
        if not (diff<1e-6): wl_mis.append((c,diff))
    # monotonic
    mono=bool(np.all(np.diff(wl[ok])>0))
    # 3) ranges
    P=A[:,1:11:2]; D=A[:,2:11:2]
    p_bad=int(np.sum((P<0)|(P>90))); d_bad=int(np.sum(np.abs(D)>360))
    # 4) cross-block consistency: NCS from Psi/Delta vs NCS block
    N_=A[:,23:38:3]; C_=A[:,24:38:3]; S_=A[:,25:38:3]
    pr=np.deg2rad(P); dr=np.deg2rad(D)
    Nc=np.cos(2*pr); Cc=np.sin(2*pr)*np.cos(dr); Sc=np.sin(2*pr)*np.sin(dr)
    m=np.isfinite(N_)&np.isfinite(Nc)
    ncs_err=float(np.nanmax(np.abs(np.where(m,N_-Nc,0))))+ \
            float(np.nanmax(np.abs(np.where(np.isfinite(C_)&np.isfinite(Cc),C_-Cc,0))))+ \
            float(np.nanmax(np.abs(np.where(np.isfinite(S_)&np.isfinite(Sc),S_-Sc,0))))
    # rho consistency
    Re_=A[:,12:22:2]; Im_=A[:,13:22:2]
    rho=np.tan(pr)*np.exp(1j*dr)
    m2=np.isfinite(Re_)
    rho_err=float(np.nanmax(np.abs(np.where(m2,Re_-rho.real,0))))+ \
            float(np.nanmax(np.abs(np.where(np.isfinite(Im_),Im_-rho.imag,0))))
    # depol range
    dep=A[:,45:50]
    dep_max=float(np.nanmax(np.abs(dep))) if np.isfinite(dep).any() else np.nan
    line='%-14s rows=%d nan(PD)=%d mono=%s wl_mis=%s Prange_bad=%d Dbad=%d NCSerr=%.4f RHOerr=%.4f depmax=%.1f'%(
        sh,n_ok,pd_nan,mono,wl_mis if wl_mis else 'OK',p_bad,d_bad,ncs_err,rho_err,dep_max)
    print(line)
    if pd_nan>0 or wl_mis or not mono or p_bad or d_bad or ncs_err>0.02 or rho_err>0.05:
        issues.append(sh)
    if ref_wl is None: ref_wl=wl
    else:
        if np.nanmax(np.abs(wl-ref_wl))>1e-6: print('   !! wl grid differs from first sheet')

# 5) cross-sheet duplicate Psi blocks (wrong-file paste)
print('\n-- cross-sheet duplicate check (Psi block identical) --')
keys=list(data.keys())
for i in range(len(keys)):
    for j in range(i+1,len(keys)):
        a=data[keys[i]][:,1:11]; b=data[keys[j]][:,1:11]
        m=np.isfinite(a)&np.isfinite(b)
        if m.sum()>1000 and np.nanmax(np.abs(np.where(m,a-b,0)))<1e-9:
            print('  DUPLICATE:',keys[i],'==',keys[j])
print('\nsheets with issues:',issues if issues else 'NONE')
np.savez(str(OUT_DIR / 'aghatcn_data.npz'),**{sh:data[sh] for sh in sheets})
print('saved aghatcn_data.npz')
