"""
import os, pathlib
DATA_DIR = pathlib.Path(os.environ.get('ELLIPS_DATA', './data'))
OUT_DIR = pathlib.Path(os.environ.get('ELLIPS_OUT', './out'))
SCRATCH_DIR = pathlib.Path(os.environ.get('ELLIPS_SCRATCH', './out'))

Spectroscopic Ellipsometry Fitting — B-spline model
Supports arbitrary angle sets and material types.
Stack: Air / Film / SiO2 / Si(substrate)
"""

import numpy as np
from scipy.optimize import minimize
from scipy.interpolate import make_interp_spline, interp1d, PchipInterpolator
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import warnings, csv, os
warnings.filterwarnings('ignore')

_HC = 1239.84193   # eV·nm

# ══════════════════════════════════════════════════════════════
# 1. Si optical constants
#    UV–VIS: Aspnes & Studna (1983) calibrated to n=3.879 @ 632.8nm
# ══════════════════════════════════════════════════════════════
_SI = np.array([
    [207,  0.883, 2.519], [248,  1.563, 3.554], [280,  2.300, 3.500],
    [300,  3.300, 3.300], [320,  4.500, 2.000], [335,  5.400, 0.700],
    [350,  5.580, 0.380], [370,  5.560, 0.310], [400,  5.566, 0.387],
    [420,  5.590, 0.350], [450,  5.400, 0.260], [480,  5.200, 0.180],
    [510,  4.900, 0.100], [540,  4.500, 0.045], [570,  4.200, 0.015],
    [600,  3.980, 0.004], [620,  3.900, 0.001], [633,  3.879, 0.000],
    [650,  3.850, 0.000], [700,  3.810, 0.000], [800,  3.737, 0.000],
    [900,  3.671, 0.000], [1000, 3.616, 0.000], [1200, 3.540, 0.000],
    [1500, 3.470, 0.000], [1688, 3.425, 0.000],
])
_si_n = interp1d(_SI[:,0], _SI[:,1], kind='linear', fill_value='extrapolate')
_si_k = interp1d(_SI[:,0], _SI[:,2], kind='linear', fill_value='extrapolate')
def n_Si(wl_nm):
    return (_si_n(wl_nm) + 1j * _si_k(wl_nm)).astype(complex)

# ══════════════════════════════════════════════════════════════
# 2. SiO2 — Sellmeier (Malitson 1965)
# ══════════════════════════════════════════════════════════════
def n_SiO2(wl_nm):
    lam = wl_nm / 1000.0
    n2 = (1 + 0.6961663*lam**2/(lam**2-0.0684043**2)
            + 0.4079426*lam**2/(lam**2-0.1162414**2)
            + 0.8974794*lam**2/(lam**2-9.896161**2))
    return np.sqrt(np.maximum(n2, 1.0)).astype(complex)

# ══════════════════════════════════════════════════════════════
# 3. Transfer Matrix Method — vectorized, corrected Fresnel
#    q_i = n_i·cos(θ_i)
#    s: rs = (qi-qj)/(qi+qj)
#    p: rp = (nj²·qi - ni²·qj)/(nj²·qi + ni²·qj)
#    prop: φ = k0·qi·d
# ══════════════════════════════════════════════════════════════
def _tmm(wl_nm, n_layers, d_list, theta0_deg):
    theta0 = np.deg2rad(theta0_deg)
    k0  = 2 * np.pi / wl_nm
    s0  = n_layers[0] * np.sin(theta0)
    q   = []
    for ni in n_layers:
        c = np.sqrt((ni**2 - s0**2).astype(complex))
        q.append(np.where(c.imag < 0, -c, c))

    N = len(wl_nm)
    def eye_b():
        M = np.zeros((N,2,2), dtype=complex); M[:,0,0]=1; M[:,1,1]=1; return M
    def mm(A, B): return np.einsum('nij,njk->nik', A, B)

    def Ip(ni, nj, qi, qj):
        d = nj**2*qi + ni**2*qj
        r = (nj**2*qi - ni**2*qj) / d
        t = 2*ni*nj*qi / d
        M = np.zeros((N,2,2), dtype=complex)
        M[:,0,0]=1/t; M[:,0,1]=r/t; M[:,1,0]=r/t; M[:,1,1]=1/t
        return M

    def Is(ni, nj, qi, qj):
        d = qi + qj
        r = (qi - qj) / d
        t = 2*qi / d
        M = np.zeros((N,2,2), dtype=complex)
        M[:,0,0]=1/t; M[:,0,1]=r/t; M[:,1,0]=r/t; M[:,1,1]=1/t
        return M

    def Pr(qi, d):
        phi = k0 * qi * d
        M = np.zeros((N,2,2), dtype=complex)
        M[:,0,0]=np.exp(1j*phi); M[:,1,1]=np.exp(-1j*phi)
        return M

    Mp = eye_b(); Ms = eye_b()
    for i, d in enumerate(d_list):
        ni,nj,qi,qj = n_layers[i],n_layers[i+1],q[i],q[i+1]
        Mp = mm(mm(Mp, Ip(ni,nj,qi,qj)), Pr(qj, d))
        Ms = mm(mm(Ms, Is(ni,nj,qi,qj)), Pr(qj, d))
    ni,nj,qi,qj = n_layers[-2],n_layers[-1],q[-2],q[-1]
    Mp = mm(Mp, Ip(ni,nj,qi,qj))
    Ms = mm(Ms, Is(ni,nj,qi,qj))
    return Mp[:,1,0]/Mp[:,0,0], Ms[:,1,0]/Ms[:,0,0]

def bruggeman_ema50(N_a, N_b):
    """Bruggeman EMA for 50% N_a + 50% N_b mixture.
    Solves: 4ε² - ε(εa+εb) - 2εa·εb = 0  → positive root.
    """
    ea = N_a**2; eb = N_b**2
    disc = ea**2 + 34*ea*eb + eb**2
    ee   = ((ea + eb) + np.sqrt(disc.astype(complex))) / 8.0
    N    = np.sqrt(ee)
    return np.where(N.imag < 0, -N, N)


def bruggeman_ema(N_a, N_b, f_a):
    """Bruggeman EMA for volume fraction f_a of N_a and (1-f_a) of N_b.
    Solves f_a(εa-ε)/(εa+2ε) + f_b(εb-ε)/(εb+2ε) = 0  →  2ε² - bε - εa·εb = 0
    with b = (2f_a - f_b)εa + (2f_b - f_a)εb ;  physical root has Im(N) ≥ 0.
    """
    ea = N_a**2; eb = N_b**2
    fb = 1.0 - f_a
    b  = (2*f_a - fb)*ea + (2*fb - f_a)*eb
    s  = np.sqrt((b**2 + 8*ea*eb).astype(complex))
    e1 = (b + s) / 4.0
    e2 = (b - s) / 4.0
    N1 = np.sqrt(e1); N1 = np.where(N1.imag < 0, -N1, N1)
    N2 = np.sqrt(e2); N2 = np.where(N2.imag < 0, -N2, N2)
    # pick the root with the larger real index (physical dense-medium branch)
    return np.where(N1.real >= N2.real, N1, N2)

def calc_psi_delta(wl, d_SiO2, d_film, N_film, angles, d_rough=0.0):
    """Returns psi (N_wl, N_ang), delta (N_wl, N_ang) in degrees.
    d_rough > 0 adds a Bruggeman EMA (50% film + 50% air) roughness layer on top.
    Stack: Air / [EMA(d_rough)] / Film(d_film) / SiO2(d_SiO2) / Si
    """
    n_amb = np.ones(len(wl), dtype=complex)
    if d_rough > 1e-6:
        N_rough = bruggeman_ema50(n_amb, N_film)
        layers  = [n_amb, N_rough, N_film, n_SiO2(wl), n_Si(wl)]
        d_list  = [d_rough, d_film, d_SiO2]
    else:
        layers = [n_amb, N_film, n_SiO2(wl), n_Si(wl)]
        d_list = [d_film, d_SiO2]
    psi_out, del_out = [], []
    for ang in angles:
        rp, rs = _tmm(wl, layers, d_list, ang)
        rho = rp / rs
        psi_out.append(np.degrees(np.arctan(np.abs(rho))))
        del_out.append(np.degrees(np.angle(rho)))
    return np.array(psi_out).T, np.array(del_out).T

# ══════════════════════════════════════════════════════════════
# 4. Data loaders
# ══════════════════════════════════════════════════════════════
def load_data(filepath, angles_deg, wl_min, wl_max):
    """Tab-delimited text: skip 2 header lines; cols λ|psi1|del1|psi2|del2|..."""
    n_ang   = len(angles_deg)
    n_cols  = 1 + 2 * n_ang
    rows = []
    with open(filepath, encoding='utf-8', errors='ignore') as f:
        for line in f.readlines()[2:]:
            cols = [c for c in line.strip().split('\t') if c.strip()]
            if len(cols) >= n_cols:
                try:
                    v = [float(c) for c in cols[:n_cols]]
                    if wl_min <= v[0] <= wl_max:
                        rows.append(v)
                except ValueError:
                    pass
    arr = np.array(rows)
    wl    = arr[:, 0]
    psi_m = arr[:, 1::2][:, :n_ang]
    del_m = arr[:, 2::2][:, :n_ang]
    return wl, psi_m, del_m


def load_data_xlsx(filepath, angles_deg, wl_min, wl_max):
    """VASE xlsx loader — auto-detects column layout from the header row.

    Handles two J.A. Woollam export styles:
      • interleaved : wl | psi1 del1 psi2 del2 ...   (Delta col = Psi col + 1)
      • grouped     : wl | psi1 psi2 psi3 | del1 del2 del3  (Delta block after Psi block)
    Detection: find the header row containing 'Wavelength', then locate the
    columns whose header contains 'Psi' and 'Delta'.
    """
    import openpyxl
    n_ang = len(angles_deg)
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    all_rows = [row for row in ws.iter_rows(values_only=True)]
    wb.close()

    # locate header row + key columns
    wl_col = psi_col = del_col = None
    hdr_idx = None
    for i, row in enumerate(all_rows[:15]):
        for j, cell in enumerate(row):
            if isinstance(cell, str):
                cl = cell.lower()
                if 'wavelength' in cl and wl_col is None:
                    wl_col = j; hdr_idx = i
                elif cl.startswith('psi') and psi_col is None:
                    psi_col = j
                elif cl.startswith('delta') and del_col is None:
                    del_col = j
        if wl_col is not None and psi_col is not None and del_col is not None:
            break

    # fallbacks if headers not found: assume interleaved with wl at col C (idx2)
    if wl_col is None:
        wl_col, psi_col, del_col, hdr_idx = 2, 3, 4, 2

    grouped = (del_col - psi_col) == n_ang   # grouped vs interleaved

    def psi_j(j): return psi_col + j if grouped else psi_col + 2*j
    def del_j(j): return del_col + j if grouped else del_col + 2*j

    rows = []
    for row in all_rows[hdr_idx + 1:]:
        if wl_col >= len(row):
            continue
        wl_val = row[wl_col]
        if wl_val is None or not isinstance(wl_val, (int, float)):
            continue
        if not (wl_min <= wl_val <= wl_max):
            continue
        vals = [float(wl_val)]
        ok = True
        for j in range(n_ang):
            pj, dj = psi_j(j), del_j(j)
            if pj >= len(row) or dj >= len(row):
                ok = False; break
            psi, dlt = row[pj], row[dj]
            if psi is None or dlt is None:
                ok = False; break
            vals.extend([float(psi), float(dlt)])
        if ok:
            rows.append(vals)

    arr   = np.array(rows)
    wl    = arr[:, 0]
    psi_m = arr[:, 1::2][:, :n_ang]
    del_m = arr[:, 2::2][:, :n_ang]
    return wl, psi_m, del_m


def _load(filepath, angles_deg, wl_min, wl_max):
    """Auto-dispatch to xlsx or text loader based on file extension."""
    if filepath.lower().endswith('.xlsx') or filepath.lower().endswith('.xls'):
        return load_data_xlsx(filepath, angles_deg, wl_min, wl_max)
    return load_data(filepath, angles_deg, wl_min, wl_max)

# ══════════════════════════════════════════════════════════════
# 5. B-spline fit — general material
# ══════════════════════════════════════════════════════════════
N_CTRL = 30

def fit(filepath, angles_deg,
        wl_min=350, wl_max=900,
        d_film_range=(5, 80),
        d_sio2_range=(0.1, 5.0),
        n_range=(0.02, 8.0),
        k_range=(1e-4, 14.0),
        n_init_flat=None,       # scalar — starting n (flat)
        k_init_flat=None,       # scalar — starting k (flat)
        d_film_init=None,
        d_sio2_init=1.0,
        reg=5e-4,
        label='Film'):

    wl, psi_m, del_m = load_data(filepath, angles_deg, wl_min, wl_max)
    n_ang  = len(angles_deg)

    # Control wavelengths (extend 50 nm beyond range for spline edge stability)
    wl_ctrl = np.linspace(wl_min - 50, wl_max + 50, N_CTRL)

    # ── Initial n, k at control points ──────────────────────
    n0 = np.full(N_CTRL, n_init_flat if n_init_flat else 1.5)
    k0 = np.full(N_CTRL, k_init_flat if k_init_flat else 0.1)
    lk0 = np.log(np.clip(k0, 1e-6, None))
    d_mg_init = d_film_init if d_film_init else np.mean(d_film_range)

    # ── Pack/unpack ──────────────────────────────────────────
    def pack(d_sio2, d_film, nc, lkc):
        return np.concatenate([[d_sio2, d_film], nc, lkc])

    def unpack(x):
        return x[0], x[1], x[2:2+N_CTRL], x[2+N_CTRL:2+2*N_CTRL]

    x0 = pack(d_sio2_init, d_mg_init, n0, lk0)

    # ── Bounds ────────────────────────────────────────────────
    lo = ([d_sio2_range[0], d_film_range[0]]
          + [n_range[0]]*N_CTRL
          + [np.log(k_range[0])]*N_CTRL)
    hi = ([d_sio2_range[1], d_film_range[1]]
          + [n_range[1]]*N_CTRL
          + [np.log(k_range[1])]*N_CTRL)
    bounds = list(zip(lo, hi))

    # ── Residual ──────────────────────────────────────────────
    def residuals(x):
        d_sio2, d_film, nc, lkc = unpack(x)
        try:
            sn  = make_interp_spline(wl_ctrl, nc,  k=3)
            sk  = make_interp_spline(wl_ctrl, lkc, k=3)
            N_film = (sn(wl) + 1j*np.exp(sk(wl))).astype(complex)
            psi_c, del_c = calc_psi_delta(wl, d_sio2, d_film, N_film, angles_deg)
        except Exception:
            return 1e10

        dpsi   = psi_c - psi_m                            # (N, n_ang)
        # Circular distance for delta (handles signed/unsigned mix)
        ddelta = (del_c - del_m + 180) % 360 - 180

        data_fit = np.mean(dpsi**2) + np.mean(ddelta**2)
        roughness = reg * (np.sum(np.diff(nc,  2)**2)
                         + np.sum(np.diff(lkc, 2)**2))
        return data_fit + roughness

    # ── Stage 1: grid search over thickness ──────────────────
    print(f"Stage 1 – thickness grid search...")
    d_vals = np.linspace(d_film_range[0], d_film_range[1], 10)
    best_x, best_f = x0.copy(), 1e10
    for dv in d_vals:
        xt = x0.copy(); xt[1] = dv
        f  = residuals(xt)
        if f < best_f:
            best_f, best_x = f, xt.copy()
    print(f"  Best start: d_SiO2={best_x[0]:.1f} nm, "
          f"d_film={best_x[1]:.1f} nm, f={best_f:.4f}")

    # ── Stage 2: L-BFGS-B optimisation ───────────────────────
    print("Stage 2 – L-BFGS-B optimisation...")
    result = minimize(residuals, best_x, method='L-BFGS-B', bounds=bounds,
                      options={'maxiter': 30000, 'ftol': 1e-14, 'gtol': 1e-9})

    x = result.x
    d_sio2, d_film, nc, lkc = unpack(x)

    sn = make_interp_spline(wl_ctrl, nc,  k=3)
    sk = make_interp_spline(wl_ctrl, lkc, k=3)
    N_film_fit = (sn(wl) + 1j*np.exp(sk(wl))).astype(complex)

    psi_c, del_c = calc_psi_delta(wl, d_sio2, d_film, N_film_fit, angles_deg)

    # ── RMSE breakdown ────────────────────────────────────────
    print(f"\n{'─'*56}")
    print(f"  SiO2 thickness : {d_sio2:.3f} nm")
    print(f"  {label} thickness : {d_film:.3f} nm")
    print(f"{'─'*56}")
    for i, ang in enumerate(angles_deg):
        rp = np.sqrt(np.mean((psi_c[:,i] - psi_m[:,i])**2))
        dd = (del_c[:,i] - del_m[:,i] + 180) % 360 - 180
        rd = np.sqrt(np.mean(dd**2))
        print(f"  {ang:5.1f}°   RMSE Ψ = {rp:.3f}°    RMSE Δ = {rd:.3f}°")
    rp_all = np.sqrt(np.mean((psi_c - psi_m)**2))
    dd_all = (del_c - del_m + 180) % 360 - 180
    rd_all = np.sqrt(np.mean(dd_all**2))
    print(f"{'─'*56}")
    print(f"  Overall         RMSE Ψ = {rp_all:.3f}°    RMSE Δ = {rd_all:.3f}°")
    print(f"{'─'*56}\n")

    return dict(wl=wl, psi_m=psi_m, del_m=del_m,
                psi_c=psi_c, del_c=del_c,
                d_sio2=d_sio2, d_film=d_film,
                N_film=N_film_fit,
                angles=angles_deg, label=label)

# ══════════════════════════════════════════════════════════════
# Helper: align calculated delta to measured convention and unwrap
# ══════════════════════════════════════════════════════════════
def _align_delta(del_c_raw, del_m_raw):
    """
    Return (del_c_plot, del_m_plot) in a consistent unwrapped convention
    so calculated and measured curves overlap visually.
    Both inputs shape (N_wl, N_ang).
    """
    N, nang = del_m_raw.shape
    del_m_plot = np.empty_like(del_m_raw)
    del_c_plot = np.empty_like(del_c_raw)

    for i in range(nang):
        # Unwrap measured (handles 0/360 wrap in raw data)
        dm = np.degrees(np.unwrap(np.radians(del_m_raw[:, i])))
        # calc_psi_delta returns signed (-180,180]; convert to continuous
        dc_raw = del_c_raw[:, i]
        # Find shift (multiple of 360) to place dc close to dm at the start
        diff0   = dm[0] - dc_raw[0]
        shift   = round(diff0 / 360.0) * 360.0
        dc      = np.degrees(np.unwrap(np.radians(dc_raw + shift)))
        # Re-anchor after unwrap in case unwrap shifted things
        re_shift = round((dm[0] - dc[0]) / 360.0) * 360.0
        dc += re_shift
        del_m_plot[:, i] = dm
        del_c_plot[:, i] = dc

    return del_c_plot, del_m_plot

# ══════════════════════════════════════════════════════════════
# 6. Cauchy + Urbach-tail + EMA roughness model
#
#  Stack : Air / EMA-rough(d_rough) / Film(d_film) / SiO2(d_sio2) / Si
#  n(λ)  = A + B/λ² + C/λ⁴
#  k(λ)  = k0 · exp(−β·(λ−λ0))    (UV absorption tail, λ0 = wl_min)
#  EMA   = Bruggeman 50% film + 50% air
#
#  Free parameters (9):
#    d_sio2, d_film, d_rough, A, B, C, k0, beta  (+C often → 0)
# ══════════════════════════════════════════════════════════════
def fit_cauchy(filepath, angles_deg,
               wl_min=400, wl_max=900,
               d_film_range=(35, 70),
               d_sio2_range=(0.5, 4.0),
               d_rough_max=5.0,
               d_film_init=50.0, d_sio2_init=1.5,
               cauchy_A_init=1.72, cauchy_B_init=30000.0,
               k0_init=0.02, beta_init=0.015,
               label='Film'):

    wl, psi_m, del_m = _load(filepath, angles_deg, wl_min, wl_max)
    n_ang = len(angles_deg)
    lam0  = wl_min

    def nk_film(wl_nm, A, B, C, k0, beta):
        n = A + B / wl_nm**2 + C / wl_nm**4
        k = k0 * np.exp(-beta * (wl_nm - lam0))
        return (n + 1j * k).astype(complex)

    # x = [d_sio2, d_film, d_rough, A, B, C, k0, beta]
    x0 = np.array([d_sio2_init, d_film_init, 1.0,
                   cauchy_A_init, cauchy_B_init, 0.0, k0_init, beta_init])

    bounds_lo = [d_sio2_range[0], d_film_range[0], 0.0,  1.2,   0,    -5e8, 0,    0.001]
    bounds_hi = [d_sio2_range[1], d_film_range[1], d_rough_max, 2.6, 1e6, 5e8, 0.5, 0.15]

    def residuals(x):
        d_sio2, d_film, d_rough, A, B, C, k0, beta = x
        try:
            N_film = nk_film(wl, A, B, C, k0, beta)
            psi_c, del_c = calc_psi_delta(wl, d_sio2, d_film, N_film,
                                           angles_deg, d_rough=d_rough)
        except Exception:
            return 1e10
        dpsi   = psi_c  - psi_m
        ddelta = (del_c - del_m + 180) % 360 - 180
        return np.mean(dpsi**2) + np.mean(ddelta**2)

    # ── Stage 1: thickness grid search ───────────────────────
    print("Stage 1 – thickness grid search (Cauchy+Urbach+rough)...")
    best_x, best_f = x0.copy(), 1e10
    for d_sio2 in np.linspace(d_sio2_range[0], d_sio2_range[1], 4):
        for d_film in np.linspace(d_film_range[0], d_film_range[1], 8):
            for d_rough in [0.0, 1.0, 2.0]:
                xt = x0.copy()
                xt[0]=d_sio2; xt[1]=d_film; xt[2]=d_rough
                f  = residuals(xt)
                if f < best_f:
                    best_f, best_x = f, xt.copy()
    print(f"  Best: d_SiO2={best_x[0]:.1f}  d_film={best_x[1]:.1f}"
          f"  d_rough={best_x[2]:.1f}  f={best_f:.4f}")

    # ── Stage 2: full optimisation ────────────────────────────
    print("Stage 2 – L-BFGS-B (Cauchy+Urbach+rough)...")
    result = minimize(residuals, best_x, method='L-BFGS-B',
                      bounds=list(zip(bounds_lo, bounds_hi)),
                      options={'maxiter': 30000, 'ftol': 1e-15, 'gtol': 1e-10})
    x = result.x
    d_sio2, d_film, d_rough, A, B, C, k0, beta = x

    N_film_fit = nk_film(wl, A, B, C, k0, beta)
    psi_c, del_c = calc_psi_delta(wl, d_sio2, d_film, N_film_fit,
                                   angles_deg, d_rough=d_rough)

    # ── RMSE report ───────────────────────────────────────────
    print(f"\n{'─'*64}")
    print(f"  Model   : Cauchy + Urbach tail + EMA roughness")
    print(f"  SiO2    : {d_sio2:.3f} nm")
    print(f"  {label:8s}: {d_film:.3f} nm  (bulk)")
    print(f"  Roughness: {d_rough:.3f} nm  (EMA top layer, 50% film + 50% air)")
    print(f"  Total    : {d_film+d_rough:.3f} nm")
    print(f"  A={A:.5f}  B={B:.1f} nm²  C={C:.3e} nm⁴")
    print(f"  k0={k0:.5f}  β={beta:.5f} nm⁻¹")
    print(f"  n(500nm)={A+B/500**2+C/500**4:.4f}  "
          f"k(400nm)={k0:.5f}  k(500nm)={k0*np.exp(-beta*100):.5f}")
    print(f"{'─'*64}")
    for i, ang in enumerate(angles_deg):
        rp = np.sqrt(np.mean((psi_c[:,i]-psi_m[:,i])**2))
        dd = (del_c[:,i]-del_m[:,i]+180)%360-180
        rd = np.sqrt(np.mean(dd**2))
        print(f"  {ang:5.1f}°   RMSE Ψ = {rp:.3f}°    RMSE Δ = {rd:.3f}°")
    dd_all = (del_c-del_m+180)%360-180
    rp_all = np.sqrt(np.mean((psi_c-psi_m)**2))
    rd_all = np.sqrt(np.mean(dd_all**2))
    mse    = np.mean((psi_c-psi_m)**2) + np.mean(dd_all**2)
    print(f"{'─'*64}")
    print(f"  Overall   RMSE Ψ = {rp_all:.3f}°    RMSE Δ = {rd_all:.3f}°")
    print(f"  MSE (Ψ²+Δ²)     = {mse:.4f}  →  √MSE = {np.sqrt(mse):.4f}°")
    print(f"{'─'*64}\n")

    return dict(wl=wl, psi_m=psi_m, del_m=del_m,
                psi_c=psi_c, del_c=del_c,
                d_sio2=d_sio2, d_film=d_film, d_rough=d_rough,
                N_film=N_film_fit,
                angles=angles_deg, label=label,
                model='Cauchy+Urbach+Rough',
                A=A, B=B, C=C, k0=k0, beta=beta)

# ══════════════════════════════════════════════════════════════
# 7. Plot & save
# ══════════════════════════════════════════════════════════════
def plot_results(r, out_png, out_csv=None):
    wl     = r['wl'];   psi_m = r['psi_m']; del_m = r['del_m']
    psi_c  = r['psi_c']; del_c = r['del_c']
    N_film = r['N_film']
    d_sio2 = r['d_sio2']; d_film = r['d_film']
    angles = r['angles']; label = r['label']
    model  = r.get('model', 'B-spline')
    n_ang  = len(angles)

    cmap   = matplotlib.colormaps['tab10']
    colors = [cmap(i / 9.0) for i in range(n_ang)]

    # ── Align delta to consistent unwrapped convention ────────
    del_c_plot, del_m_plot = _align_delta(del_c, del_m)

    d_rough = r.get('d_rough', 0.0)
    if model == 'Cauchy+Urbach+Rough':
        subtitle = (f"A={r['A']:.4f}, B={r['B']:.0f} nm²,"
                    f" k₀={r['k0']:.4f}, β={r['beta']:.4f} nm⁻¹"
                    f"  |  d_rough={d_rough:.2f} nm")
    elif model in ('Cauchy+Urbach', 'Cauchy'):
        subtitle = (f"A={r['A']:.4f}, B={r['B']:.0f} nm²,"
                    f" k₀={r.get('k0',0):.4f}")
    elif model == 'Bspline-n(KK)+Urbach-k':
        subtitle = (f"k₀={r.get('k0',0):.4f}, β={r.get('beta',0):.4f} nm⁻¹"
                    f"  |  d_rough={d_rough:.2f} nm  (KK+mono)")
    else:
        subtitle = ''

    fig, axes = plt.subplots(2, 2, figsize=(14, 9))
    fig.suptitle(
        f'{label} Ellipsometry Fit ({model})\n'
        f'd(SiO₂)={d_sio2:.2f} nm  d({label})={d_film:.2f} nm  {subtitle}',
        fontsize=9, fontweight='bold')

    ax_p, ax_d = axes[0,0], axes[0,1]
    ax_n, ax_k  = axes[1,0], axes[1,1]

    for i, (ang, col) in enumerate(zip(angles, colors)):
        lab = f'{ang:.0f}°'
        # Ψ: dots = measured, line = calculated
        ax_p.plot(wl, psi_m[:,i], '.', color=col, ms=2.0, alpha=0.4)
        ax_p.plot(wl, psi_c[:,i], '-', color=col, lw=1.2, label=lab)
        # Δ: aligned convention for both
        ax_d.plot(wl, del_m_plot[:,i], '.', color=col, ms=2.0, alpha=0.4)
        ax_d.plot(wl, del_c_plot[:,i], '-', color=col, lw=1.2, label=lab)

    for ax, title, ylab in [(ax_p,'Ψ (Psi)','Ψ (deg)'), (ax_d,'Δ (Delta)','Δ (deg)')]:
        ax.set_title(title); ax.set_xlabel('Wavelength (nm)'); ax.set_ylabel(ylab)
        ax.legend(fontsize=7, ncol=2); ax.grid(True, alpha=0.2)

    ax_n.plot(wl, N_film.real, 'b-', lw=1.5)
    ax_n.set_title(f'{label}  n(λ)'); ax_n.set_xlabel('Wavelength (nm)'); ax_n.set_ylabel('n')
    ax_n.set_ylim(bottom=0); ax_n.grid(True, alpha=0.2)

    ax_k.plot(wl, N_film.imag, 'r-', lw=1.5)
    ax_k.set_title(f'{label}  k(λ)'); ax_k.set_xlabel('Wavelength (nm)'); ax_k.set_ylabel('k')
    ax_k.set_ylim(bottom=0); ax_k.grid(True, alpha=0.2)

    plt.tight_layout()
    plt.savefig(out_png, dpi=150)
    print(f"Figure → {out_png}")

    if out_csv:
        with open(out_csv, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(['wavelength_nm', 'n', 'k'])
            for wl_i, ni, ki in zip(wl, N_film.real, N_film.imag):
                w.writerow([f'{wl_i:.3f}', f'{ni:.6f}', f'{ki:.6f}'])
        print(f"n,k table → {out_csv}")

# ══════════════════════════════════════════════════════════════
# 8. Presets
# ══════════════════════════════════════════════════════════════
PRESETS = {
    'MgAg': dict(
        angles_deg=[65.0, 70.0, 75.0],
        wl_min=400, wl_max=1000,
        d_film_range=(3, 40),
        n_range=(0.02, 8.0), k_range=(0.05, 14.0),
        n_init_flat=0.5, k_init_flat=3.0,
        d_film_init=13.0, d_sio2_init=1.0,
        reg=5e-4, label='MgAg',
    ),
    'Bphen': dict(
        angles_deg=[45.0, 50.0, 55.0, 60.0, 65.0, 70.0, 75.0],
        wl_min=300, wl_max=900,
        d_film_range=(35, 70),
        d_sio2_range=(0.5, 4.0),
        n_range=(1.2, 2.8), k_range=(1e-4, 1.5),
        n_init_flat=1.75, k_init_flat=0.02,
        d_film_init=50.0, d_sio2_init=1.5,
        reg=1e-3, label='Bphen',
    ),
}

CAUCHY_PRESETS = {
    'Bphen': dict(
        angles_deg=[45.0, 50.0, 55.0, 60.0, 65.0, 70.0, 75.0],
        wl_min=400, wl_max=900,
        d_film_range=(35, 70),
        d_sio2_range=(0.5, 4.0),
        d_film_init=50.0, d_sio2_init=1.5,
        label='Bphen',
    ),
    'HATCN': dict(
        angles_deg=[65.0, 70.0, 75.0],
        wl_min=380, wl_max=900,
        d_film_range=(20, 45),
        d_sio2_range=(1.0, 3.0),
        d_film_init=27.0, d_sio2_init=1.5,
        cauchy_A_init=1.85, cauchy_B_init=28000.0,
        k0_init=0.08, beta_init=0.05,
        label='HATCN',
    ),
}

# ══════════════════════════════════════════════════════════════
# 9. B-spline n(λ) + Urbach k + EMA roughness  [KK-consistent]
#
#  Stack : Air / EMA-rough(d_rough) / Film(d_film) / SiO2 / Si
#  n(λ)  : n_bg(λ) [B-spline, N_CTRL_N ctrl pts]
#           + n_KK(λ) [KK transform of Urbach k, precomputed matrix]
#  k(λ)  : k0 · exp(−β·(λ−λ_min)) ≥ 0   (Urbach UV tail)
#
#  KK: H_kk precomputed once; n_KK = H_kk @ k_vals each residuals call.
#      n and k are KK-consistent for the modelled Urbach portion.
#
#  Physical constraints:
#    k ≥ 0             : bounds (k0 ≥ 0) + np.maximum clip
#    n monotone ↓      : soft penalty  reg_mono · Σ max(0, Δn_i)²
#    d_SiO₂ ≥ 1.0 nm  : lower bound (native oxide)
#    d_rough ≤ 5.0 nm  : upper bound
#
#  Free params: d_sio2, d_film, d_rough, n_bg_ctrl×N, k0, beta
# ══════════════════════════════════════════════════════════════
N_CTRL_N = 20   # spline knots for n_bg(λ)

def fit_bspline_n(filepath, angles_deg,
                  wl_min=400, wl_max=900,
                  d_film_range=(35, 70),
                  d_sio2_range=(1.0, 4.0),
                  d_sio2_fixed=None,        # fix SiO2 thickness (nm); None = free
                  d_rough_max=5.0,
                  d_film_init=50.0, d_sio2_init=1.5,
                  reg=5e-4, reg_mono=1.0,
                  reg_prior=0.0,            # soft prior pulling n → Cauchy literature
                  n_cauchy_A=1.71261, n_cauchy_B=34480.4,   # init & prior Cauchy
                  k0_seed=0.067, beta_seed=0.066,
                  n_lo=1.0, n_hi=2.8,
                  label='Film'):

    wl, psi_m, del_m = _load(filepath, angles_deg, wl_min, wl_max)
    N    = len(wl)
    lam0 = wl_min
    wl_ctrl = np.linspace(wl_min - 30, wl_max + 30, N_CTRL_N)

    # ── Literature n prior curve (Cauchy) ─────────────────────────
    # n_prior(λ) = n_cauchy_A + n_cauchy_B/λ² evaluated on wl grid.
    # reg_prior * mean((n - n_prior)²) added to residuals.
    n_prior_wl = n_cauchy_A + n_cauchy_B / wl**2

    # ── Precompute KK matrix ──────────────────────────────────────
    E_wl_asc = (1239.84 / wl)[::-1]
    dE_asc   = np.abs(np.gradient(E_wl_asc))
    ii       = np.arange(N)[:, None]; jj = np.arange(N)[None, :]
    Ei = E_wl_asc[ii]; Ej = E_wl_asc[jj]; dEj = dE_asc[jj]
    eps_E = 0.05
    with np.errstate(divide='ignore', invalid='ignore'):
        H_E = np.where(np.abs(Ej - Ei) > eps_E,
                       (2.0/np.pi)*Ej/(Ej**2-Ei**2)*dEj, 0.0)
    H_kk = H_E[::-1, ::-1].copy()

    # ── Initial n_bg control points (seed from Cauchy − n_KK(init k)) ──
    n0_cauchy = n_cauchy_A + n_cauchy_B / wl_ctrl**2
    k_init    = np.maximum(k0_seed * np.exp(-beta_seed*(wl-lam0)), 0.0)
    n_kk_init = H_kk @ k_init
    n_kk_ctrl = PchipInterpolator(wl, n_kk_init)(wl_ctrl)
    n0_bg     = n0_cauchy - n_kk_ctrl

    # ── Pack / unpack — two variants depending on d_sio2_fixed ───
    fix_sio2 = d_sio2_fixed is not None

    if fix_sio2:
        d_sio2_val = float(d_sio2_fixed)
        # x = [d_film, d_rough, nc×N, k0, beta]
        def pack(d_film, d_rough, nc, k0, beta):
            return np.concatenate([[d_film, d_rough], nc, [k0, beta]])
        def unpack(x):
            return d_sio2_val, x[0], x[1], x[2:2+N_CTRL_N], x[2+N_CTRL_N], x[2+N_CTRL_N+1]
        x0 = pack(d_film_init, 1.0, n0_bg, k0_seed, beta_seed)
        bounds_lo = ([d_film_range[0], 0.0] + [n_lo]*N_CTRL_N + [0.0, 0.001])
        bounds_hi = ([d_film_range[1], d_rough_max] + [n_hi]*N_CTRL_N + [0.5, 0.20])
        _film_idx = 0   # index of d_film in x for grid search
    else:
        # x = [d_sio2, d_film, d_rough, nc×N, k0, beta]
        def pack(d_sio2, d_film, d_rough, nc, k0, beta):
            return np.concatenate([[d_sio2, d_film, d_rough], nc, [k0, beta]])
        def unpack(x):
            return x[0], x[1], x[2], x[3:3+N_CTRL_N], x[3+N_CTRL_N], x[3+N_CTRL_N+1]
        x0 = pack(d_sio2_init, d_film_init, 1.0, n0_bg, k0_seed, beta_seed)
        bounds_lo = ([d_sio2_range[0], d_film_range[0], 0.0] + [n_lo]*N_CTRL_N + [0.0, 0.001])
        bounds_hi = ([d_sio2_range[1], d_film_range[1], d_rough_max] + [n_hi]*N_CTRL_N + [0.5, 0.20])
        _film_idx = 1

    # ── Residuals ─────────────────────────────────────────────────
    def residuals(x):
        d_sio2, d_film, d_rough, nc, k0, beta = unpack(x)
        try:
            k_vals = np.maximum(k0 * np.exp(-beta*(wl-lam0)), 0.0)
            n_kk   = H_kk @ k_vals
            n_bg   = make_interp_spline(wl_ctrl, nc, k=3)(wl)
            n_vals = n_bg + n_kk
            N_film = (n_vals + 1j*k_vals).astype(complex)
            psi_c, del_c = calc_psi_delta(wl, d_sio2, d_film, N_film,
                                           angles_deg, d_rough=d_rough)
        except Exception:
            return 1e10
        dpsi   = psi_c - psi_m
        ddelta = (del_c - del_m + 180) % 360 - 180
        smooth = reg * np.sum(np.diff(nc, 2)**2)
        mono   = reg_mono * np.sum(np.maximum(0.0, np.diff(nc))**2)
        prior  = reg_prior * np.mean((n_vals - n_prior_wl)**2)
        return np.mean(dpsi**2) + np.mean(ddelta**2) + smooth + mono + prior

    # ── Stage 1: grid search over d_film (and d_sio2 if free) ────
    sio2_label = f'{d_sio2_fixed:.2f}nm (fixed)' if fix_sio2 else 'free'
    print(f"Stage 1 – grid search  [d_SiO2={sio2_label}]...")
    best_x, best_f = x0.copy(), 1e10
    for d_film in np.linspace(d_film_range[0], d_film_range[1], 10):
        if fix_sio2:
            xt = x0.copy(); xt[_film_idx] = d_film
            f  = residuals(xt)
            if f < best_f:
                best_f, best_x = f, xt.copy()
        else:
            for d_sio2 in [d_sio2_range[0], 1.5, d_sio2_range[1]]:
                xt = x0.copy(); xt[0] = d_sio2; xt[_film_idx] = d_film
                f  = residuals(xt)
                if f < best_f:
                    best_f, best_x = f, xt.copy()
    _d_sio2_best, _d_film_best = unpack(best_x)[:2]
    print(f"  Best: d_SiO2={_d_sio2_best:.2f}  d_film={_d_film_best:.1f} nm  f={best_f:.4f}")

    # ── Stage 2: L-BFGS-B ────────────────────────────────────────
    prior_str = f'prior={reg_prior}' if reg_prior > 0 else 'no prior'
    print(f"Stage 2 – L-BFGS-B  [{prior_str}, mono={reg_mono}]...")
    result = minimize(residuals, best_x, method='L-BFGS-B',
                      bounds=list(zip(bounds_lo, bounds_hi)),
                      options={'maxiter': 50000, 'ftol': 1e-15, 'gtol': 1e-10})
    x = result.x
    d_sio2, d_film, d_rough, nc, k0, beta = unpack(x)

    k_vals     = np.maximum(k0 * np.exp(-beta*(wl-lam0)), 0.0)
    n_kk       = H_kk @ k_vals
    n_bg       = make_interp_spline(wl_ctrl, nc, k=3)(wl)
    n_vals     = n_bg + n_kk
    N_film_fit = (n_vals + 1j*k_vals).astype(complex)

    psi_c, del_c = calc_psi_delta(wl, d_sio2, d_film, N_film_fit,
                                   angles_deg, d_rough=d_rough)

    # ── Report ────────────────────────────────────────────────────
    dd_all = (del_c - del_m + 180) % 360 - 180
    rp_all = np.sqrt(np.mean((psi_c - psi_m)**2))
    rd_all = np.sqrt(np.mean(dd_all**2))
    mse    = np.mean((psi_c - psi_m)**2) + np.mean(dd_all**2)
    n_dev  = np.sqrt(np.mean((n_vals - n_prior_wl)**2))

    def _n_at(nm):
        return float(n_vals[np.argmin(np.abs(wl - nm))])

    dn_max  = float(np.max(np.diff(n_vals)))
    mono_ok = dn_max <= 2e-3

    sio2_str = (f'{d_sio2:.3f} nm  (FIXED)'
                if fix_sio2 else
                f'{d_sio2:.3f} nm  (bound ≥{d_sio2_range[0]:.1f} nm)')
    print(f"\n{'─'*72}")
    print(f"  Model    : B-spline n_bg(KK) + Urbach k + EMA rough  [{prior_str}]")
    print(f"  SiO2     : {sio2_str}")
    print(f"  {label:8s} : {d_film:.3f} nm (bulk) + {d_rough:.3f} nm (rough) "
          f"= {d_film+d_rough:.3f} nm total")
    print(f"  k0={k0:.5f}  β={beta:.5f} nm⁻¹   k_min={k_vals.min():.2e}  (k≥0 ✓)")
    print(f"  n(400nm)={_n_at(400):.4f}  n(500nm)={_n_at(500):.4f}  "
          f"n(600nm)={_n_at(600):.4f}  n(633nm)={_n_at(633):.4f}  n(900nm)={_n_at(900):.4f}")
    print(f"  n vs prior (Cauchy A={n_cauchy_A:.4f} B={n_cauchy_B:.0f}):  RMS dev = {n_dev:.4f}")
    print(f"  n monotone ↓: {'✓' if mono_ok else f'✗ (max local +{dn_max:.4f})'}")
    print(f"{'─'*72}")
    for i, ang in enumerate(angles_deg):
        rp = np.sqrt(np.mean((psi_c[:,i]-psi_m[:,i])**2))
        dd = (del_c[:,i]-del_m[:,i]+180)%360-180
        rd = np.sqrt(np.mean(dd**2))
        print(f"  {ang:5.1f}°   RMSE Ψ = {rp:.3f}°    RMSE Δ = {rd:.3f}°")
    print(f"{'─'*72}")
    print(f"  Overall   RMSE Ψ = {rp_all:.3f}°    RMSE Δ = {rd_all:.3f}°")
    print(f"  MSE (Ψ²+Δ²)     = {mse:.4f}  →  √MSE = {np.sqrt(mse):.4f}°")
    print(f"{'─'*72}\n")

    return dict(wl=wl, psi_m=psi_m, del_m=del_m,
                psi_c=psi_c, del_c=del_c,
                d_sio2=d_sio2, d_film=d_film, d_rough=d_rough,
                N_film=N_film_fit,
                angles=angles_deg, label=label,
                model='Bspline-n(KK)+Urbach-k',
                k0=k0, beta=beta)

BSPLINE_N_PRESETS = {
    'Bphen': dict(
        angles_deg=[45.0, 50.0, 55.0, 60.0, 65.0, 70.0, 75.0],
        wl_min=400, wl_max=900,
        d_film_range=(35, 70),
        d_sio2_range=(1.0, 4.0),
        d_rough_max=5.0,
        d_film_init=49.5, d_sio2_init=1.5,
        reg=5e-4, reg_mono=5000,
        label='Bphen',
    ),
    'HATCN': dict(
        # Literature: n(600nm)=1.78 (US Pat.10892418), bandgap 3.3eV→k≈0 above ~430nm
        # Cauchy A=1.740, B=14400 nm² → n_prior(600nm)=1.780
        # Data-consistent n≈1.99 at 600nm; reg_prior=10 is best data/prior tradeoff
        angles_deg=[65.0, 70.0, 75.0],
        wl_min=380, wl_max=900,
        d_film_range=(20, 45),
        d_sio2_fixed=1.5,            # native oxide fixed (d_SiO2 free → 3.5nm artifact)
        d_rough_max=5.0,
        d_film_init=27.0, d_sio2_init=1.5,
        reg=5e-4, reg_mono=5000,
        reg_prior=10.0,              # gentle pull toward literature n(600nm)=1.78
        n_cauchy_A=1.740, n_cauchy_B=14400.0,
        k0_seed=0.22, beta_seed=0.037,
        n_lo=1.3, n_hi=2.2,
        label='HATCN',
    ),
}

# ══════════════════════════════════════════════════════════════
# 10. Uniaxial anisotropic model  (no, ne)
#
#  Stack : Air / Film(no, ne, d_film) / SiO2(d_sio2) / Si
#  Optical axis: z (⊥ substrate)   →  no = in-plane, ne = out-of-plane
#
#  s-pol sees only no; p-pol sees both through:
#    qe = sqrt(no² − no²·s²/ne²)
#  Interface Fresnel for p-pol at iso(ni,qi)↔film(no,ne,qe):
#    rp = (no²·qi − ni²·qe) / (no²·qi + ni²·qe)
#  → equivalent to Ip(ni, no, qi, qe) in our existing TMM code.
#
#  Free params: d_film, nc_o×N, nc_e×N, k0, beta
#  d_sio2 fixed (native oxide)
# ══════════════════════════════════════════════════════════════

def _tmm_uniaxial(wl_nm, n_amb, no_f, ne_f, n_sio2, n_si, d_film, d_sio2, theta0_deg):
    """TMM for Air/Film(uniaxial,no,ne)/SiO2/Si.
    no_f, ne_f: complex arrays (N_wl,). Returns (rp, rs)."""
    theta0 = np.deg2rad(theta0_deg)
    k0  = 2*np.pi/wl_nm
    N   = len(wl_nm)
    s2  = (n_amb*np.sin(theta0))**2

    def q_iso(n):
        c = np.sqrt((n**2 - s2).astype(complex))
        return np.where(c.imag < 0, -c, c)

    q0  = q_iso(n_amb)
    qo  = q_iso(no_f)                                # ordinary (s-pol)
    no2 = no_f**2; ne2 = ne_f**2
    qe  = np.sqrt((no2 - no2*s2/ne2).astype(complex))  # extraordinary (p-pol)
    qe  = np.where(qe.imag < 0, -qe, qe)
    q2  = q_iso(n_sio2)
    q3  = q_iso(n_si)

    def eye_b():
        M = np.zeros((N,2,2),dtype=complex); M[:,0,0]=1; M[:,1,1]=1; return M
    def mm(A,B): return np.einsum('nij,njk->nik',A,B)

    def Ip(ni, nj, qi, qj):              # p-pol interface matrix (iso or aniso)
        d = nj**2*qi + ni**2*qj
        r = (nj**2*qi - ni**2*qj)/d
        t = 2*ni*nj*qi/d
        M = np.zeros((N,2,2),dtype=complex)
        M[:,0,0]=1/t; M[:,0,1]=r/t; M[:,1,0]=r/t; M[:,1,1]=1/t
        return M

    def Is(ni, nj, qi, qj):             # s-pol interface matrix
        d = qi + qj
        r = (qi-qj)/d; t = 2*qi/d
        M = np.zeros((N,2,2),dtype=complex)
        M[:,0,0]=1/t; M[:,0,1]=r/t; M[:,1,0]=r/t; M[:,1,1]=1/t
        return M

    def Pr(qi, d):
        phi = k0*qi*d
        M = np.zeros((N,2,2),dtype=complex)
        M[:,0,0]=np.exp(1j*phi); M[:,1,1]=np.exp(-1j*phi)
        return M

    # p-pol: extraordinary wavevector qe in film; interface uses no_f as effective nj
    Mp = eye_b()
    Mp = mm(mm(Mp, Ip(n_amb, no_f, q0, qe)), Pr(qe, d_film))
    Mp = mm(mm(Mp, Ip(no_f, n_sio2, qe, q2)), Pr(q2, d_sio2))
    Mp = mm(Mp,     Ip(n_sio2, n_si, q2, q3))

    # s-pol: ordinary wavevector qo in film
    Ms = eye_b()
    Ms = mm(mm(Ms, Is(n_amb, no_f, q0, qo)), Pr(qo, d_film))
    Ms = mm(mm(Ms, Is(no_f, n_sio2, qo, q2)), Pr(q2, d_sio2))
    Ms = mm(Ms,     Is(n_sio2, n_si, q2, q3))

    return Mp[:,1,0]/Mp[:,0,0], Ms[:,1,0]/Ms[:,0,0]


def calc_psi_delta_aniso(wl, d_sio2, d_film, no_film, ne_film, angles):
    """Psi/Delta for uniaxial film. Returns (psi, delta) each (N_wl, N_ang)."""
    n_amb = np.ones(len(wl), dtype=complex)
    psi_out, del_out = [], []
    for ang in angles:
        rp, rs = _tmm_uniaxial(wl, n_amb, no_film, ne_film,
                                n_SiO2(wl), n_Si(wl), d_film, d_sio2, ang)
        rho = rp/rs
        psi_out.append(np.degrees(np.arctan(np.abs(rho))))
        del_out.append(np.degrees(np.angle(rho)))
    return np.array(psi_out).T, np.array(del_out).T


N_CTRL_A = 15   # ctrl pts per polarization for anisotropic fit

def fit_bspline_aniso(filepath, angles_deg,
                      wl_min=380, wl_max=900,
                      d_film_range=(20, 45),
                      d_sio2_fixed=1.5,
                      d_film_init=28.0,
                      reg=5e-4, reg_mono=3000,
                      reg_prior_o=5.0, reg_prior_e=5.0,
                      n_cauchy_Ao=1.960, n_cauchy_Bo=18000.0,   # no init/prior
                      n_cauchy_Ae=1.740, n_cauchy_Be=14400.0,   # ne init/prior
                      k0_seed=0.22, beta_seed=0.037,
                      n_lo=1.2, n_hi=2.4,
                      label='Film'):
    """B-spline fit with uniaxial anisotropy (no, ne separate).
    d_sio2 is fixed. Fits: d_film, nc_o×N, nc_e×N, k0, beta."""

    wl, psi_m, del_m = _load(filepath, angles_deg, wl_min, wl_max)
    N = len(wl); lam0 = wl_min
    wl_ctrl = np.linspace(wl_min - 30, wl_max + 30, N_CTRL_A)
    d_sio2  = float(d_sio2_fixed)

    # ── KK matrix ────────────────────────────────────────────────
    E_wl_asc = (1239.84/wl)[::-1]
    dE_asc   = np.abs(np.gradient(E_wl_asc))
    ii = np.arange(N)[:,None]; jj = np.arange(N)[None,:]
    Ei = E_wl_asc[ii]; Ej = E_wl_asc[jj]; dEj = dE_asc[jj]
    with np.errstate(divide='ignore', invalid='ignore'):
        H_E = np.where(np.abs(Ej-Ei) > 0.05,
                       (2.0/np.pi)*Ej/(Ej**2-Ei**2)*dEj, 0.0)
    H_kk = H_E[::-1,::-1].copy()

    # ── Prior curves ──────────────────────────────────────────────
    n_prior_o = n_cauchy_Ao + n_cauchy_Bo/wl**2
    n_prior_e = n_cauchy_Ae + n_cauchy_Be/wl**2

    # ── Initial ctrl pts ─────────────────────────────────────────
    k_init   = np.maximum(k0_seed*np.exp(-beta_seed*(wl-lam0)), 0.0)
    n_kk_0   = H_kk @ k_init
    nkk_ctrl = PchipInterpolator(wl, n_kk_0)(wl_ctrl)

    def init_bg(A, B):
        return (A + B/wl_ctrl**2) - nkk_ctrl

    nc_o0 = init_bg(n_cauchy_Ao, n_cauchy_Bo)
    nc_e0 = init_bg(n_cauchy_Ae, n_cauchy_Be)

    # ── Pack / unpack ─────────────────────────────────────────────
    # x = [d_film, nc_o×N, nc_e×N, k0, beta]
    M = N_CTRL_A
    def pack(d_film, nc_o, nc_e, k0, beta):
        return np.concatenate([[d_film], nc_o, nc_e, [k0, beta]])
    def unpack(x):
        return x[0], x[1:1+M], x[1+M:1+2*M], x[1+2*M], x[2+2*M]

    x0 = pack(d_film_init, nc_o0, nc_e0, k0_seed, beta_seed)

    bounds_lo = ([d_film_range[0]]
                 + [n_lo]*M + [n_lo]*M + [0.0, 0.001])
    bounds_hi = ([d_film_range[1]]
                 + [n_hi]*M + [n_hi]*M + [0.5, 0.20])

    # ── Residuals ─────────────────────────────────────────────────
    def residuals(x):
        d_film, nc_o, nc_e, k0, beta = unpack(x)
        try:
            k_vals = np.maximum(k0*np.exp(-beta*(wl-lam0)), 0.0)
            n_kk   = H_kk @ k_vals
            spl    = make_interp_spline(wl_ctrl, np.ones(M), k=3)   # dummy, reuse below
            no_bg  = make_interp_spline(wl_ctrl, nc_o, k=3)(wl)
            ne_bg  = make_interp_spline(wl_ctrl, nc_e, k=3)(wl)
            no_vals = no_bg + n_kk
            ne_vals = ne_bg + n_kk
            no_c = (no_vals + 1j*k_vals).astype(complex)
            ne_c = (ne_vals + 1j*k_vals).astype(complex)
            psi_c, del_c = calc_psi_delta_aniso(wl, d_sio2, d_film, no_c, ne_c, angles_deg)
        except Exception:
            return 1e10
        dpsi   = psi_c - psi_m
        ddelta = (del_c - del_m + 180)%360 - 180
        smooth_o = reg * np.sum(np.diff(nc_o, 2)**2)
        smooth_e = reg * np.sum(np.diff(nc_e, 2)**2)
        mono_o   = reg_mono * np.sum(np.maximum(0.0, np.diff(nc_o))**2)
        mono_e   = reg_mono * np.sum(np.maximum(0.0, np.diff(nc_e))**2)
        prior_o  = reg_prior_o * np.mean((no_vals - n_prior_o)**2)
        prior_e  = reg_prior_e * np.mean((ne_vals - n_prior_e)**2)
        return (np.mean(dpsi**2) + np.mean(ddelta**2)
                + smooth_o + smooth_e + mono_o + mono_e + prior_o + prior_e)

    # ── Stage 1: grid search ──────────────────────────────────────
    print(f"Stage 1 – grid search  [uniaxial no/ne, d_SiO2={d_sio2:.2f}nm fixed]")
    best_x, best_f = x0.copy(), 1e10
    for d_film in np.linspace(d_film_range[0], d_film_range[1], 10):
        xt = x0.copy(); xt[0] = d_film
        f  = residuals(xt)
        if f < best_f:
            best_f, best_x = f, xt.copy()
    print(f"  Best: d_film={best_x[0]:.1f} nm  f={best_f:.4f}")

    # ── Stage 2: L-BFGS-B ────────────────────────────────────────
    print(f"Stage 2 – L-BFGS-B  [{N_CTRL_A} ctrl×2 + k, uniaxial]...")
    result = minimize(residuals, best_x, method='L-BFGS-B',
                      bounds=list(zip(bounds_lo, bounds_hi)),
                      options={'maxiter': 60000, 'ftol': 1e-15, 'gtol': 1e-10})
    x = result.x
    d_film, nc_o, nc_e, k0, beta = unpack(x)

    k_vals  = np.maximum(k0*np.exp(-beta*(wl-lam0)), 0.0)
    n_kk    = H_kk @ k_vals
    no_vals = make_interp_spline(wl_ctrl, nc_o, k=3)(wl) + n_kk
    ne_vals = make_interp_spline(wl_ctrl, nc_e, k=3)(wl) + n_kk
    no_c    = (no_vals + 1j*k_vals).astype(complex)
    ne_c    = (ne_vals + 1j*k_vals).astype(complex)

    psi_c, del_c = calc_psi_delta_aniso(wl, d_sio2, d_film, no_c, ne_c, angles_deg)

    # ── Report ────────────────────────────────────────────────────
    dd_all = (del_c - del_m + 180)%360 - 180
    rp_all = np.sqrt(np.mean((psi_c - psi_m)**2))
    rd_all = np.sqrt(np.mean(dd_all**2))
    mse    = np.mean((psi_c - psi_m)**2) + np.mean(dd_all**2)

    def _at(arr, nm):
        return float(arr[np.argmin(np.abs(wl - nm))])

    print(f"\n{'─'*72}")
    print(f"  Model    : Uniaxial B-spline(KK)  no/ne + Urbach k")
    print(f"  SiO2     : {d_sio2:.3f} nm (FIXED)")
    print(f"  {label:8s} : {d_film:.3f} nm")
    print(f"  k0={k0:.5f}  β={beta:.5f} nm⁻¹   k_min={k_vals.min():.2e}")
    print(f"  no: {_at(no_vals,400):.4f}@400  {_at(no_vals,500):.4f}@500  "
          f"{_at(no_vals,600):.4f}@600  {_at(no_vals,633):.4f}@633  {_at(no_vals,900):.4f}@900 nm")
    print(f"  ne: {_at(ne_vals,400):.4f}@400  {_at(ne_vals,500):.4f}@500  "
          f"{_at(ne_vals,600):.4f}@600  {_at(ne_vals,633):.4f}@633  {_at(ne_vals,900):.4f}@900 nm")
    print(f"  birefringence Δn(633nm) = no-ne = "
          f"{_at(no_vals,633)-_at(ne_vals,633):+.4f}")
    print(f"{'─'*72}")
    for i, ang in enumerate(angles_deg):
        rp = np.sqrt(np.mean((psi_c[:,i]-psi_m[:,i])**2))
        dd = (del_c[:,i]-del_m[:,i]+180)%360-180
        rd = np.sqrt(np.mean(dd**2))
        print(f"  {ang:5.1f}°   RMSE Ψ = {rp:.3f}°    RMSE Δ = {rd:.3f}°")
    print(f"{'─'*72}")
    print(f"  Overall   RMSE Ψ = {rp_all:.3f}°    RMSE Δ = {rd_all:.3f}°")
    print(f"  MSE (Ψ²+Δ²)     = {mse:.4f}  →  √MSE = {np.sqrt(mse):.4f}°")
    print(f"{'─'*72}\n")

    return dict(wl=wl, psi_m=psi_m, del_m=del_m,
                psi_c=psi_c, del_c=del_c,
                d_sio2=d_sio2, d_film=d_film, d_rough=0.0,
                N_film=no_c,   # use no for plot (ordinary)
                no=no_vals, ne=ne_vals, k=k_vals,
                angles=angles_deg, label=label,
                model='Uniaxial-Bspline(KK)',
                k0=k0, beta=beta)


ANISO_PRESETS = {
    'HATCN': dict(
        angles_deg=[65.0, 70.0, 75.0],
        wl_min=380, wl_max=900,
        d_film_range=(20, 45),
        d_sio2_fixed=1.5,
        d_film_init=28.0,
        reg=5e-4, reg_mono=3000,
        reg_prior_o=5.0, reg_prior_e=5.0,
        # no init ≈ isotropic fit result (1.96@600nm); ne init ≈ literature (1.78@600nm)
        n_cauchy_Ao=1.920, n_cauchy_Bo=18000.0,
        n_cauchy_Ae=1.720, n_cauchy_Be=12000.0,
        k0_seed=0.22, beta_seed=0.037,
        n_lo=1.2, n_hi=2.4,
        label='HATCN',
    ),
}


def plot_results_aniso(r, out_png, out_csv=None):
    """Plot for uniaxial aniso results: no, ne, k separately."""
    wl    = r['wl']; psi_m = r['psi_m']; del_m = r['del_m']
    psi_c = r['psi_c']; del_c = r['del_c']
    no = r['no']; ne = r['ne']; k = r['k']
    angles = r['angles']; label = r['label']
    d_film = r['d_film']; d_sio2 = r['d_sio2']
    n_ang = len(angles)
    cmap = matplotlib.colormaps['tab10']
    colors = [cmap(i/9.0) for i in range(n_ang)]
    del_c_plot, del_m_plot = _align_delta(del_c, del_m)

    fig, axes = plt.subplots(2, 2, figsize=(14, 9))
    fig.suptitle(f'{label} Uniaxial Ellipsometry Fit\n'
                 f'd(SiO₂)={d_sio2:.2f} nm  d({label})={d_film:.2f} nm  '
                 f'(no/ne B-spline+KK+Urbach-k)',
                 fontsize=9, fontweight='bold')
    ax_p, ax_d, ax_n, ax_k = axes[0,0], axes[0,1], axes[1,0], axes[1,1]

    for i, (ang, col) in enumerate(zip(angles, colors)):
        lab = f'{ang:.0f}°'
        ax_p.plot(wl, psi_m[:,i],  '.', color=col, ms=2, alpha=0.4)
        ax_p.plot(wl, psi_c[:,i],  '-', color=col, lw=1.2, label=lab)
        ax_d.plot(wl, del_m_plot[:,i], '.', color=col, ms=2, alpha=0.4)
        ax_d.plot(wl, del_c_plot[:,i], '-', color=col, lw=1.2, label=lab)
    for ax, title, ylab in [(ax_p,'Ψ (Psi)','Ψ (deg)'), (ax_d,'Δ (Delta)','Δ (deg)')]:
        ax.set_title(title); ax.set_xlabel('Wavelength (nm)'); ax.set_ylabel(ylab)
        ax.legend(fontsize=7, ncol=2); ax.grid(True, alpha=0.2)

    ax_n.plot(wl, no, 'b-', lw=1.5, label='no (in-plane)')
    ax_n.plot(wl, ne, 'r--', lw=1.5, label='ne (out-of-plane)')
    ax_n.set_title(f'{label}  n(λ)  — uniaxial')
    ax_n.set_xlabel('Wavelength (nm)'); ax_n.set_ylabel('n')
    ax_n.set_ylim(bottom=0); ax_n.legend(); ax_n.grid(True, alpha=0.2)

    ax_k.plot(wl, k, 'r-', lw=1.5, label='k (shared)')
    ax_k.set_title(f'{label}  k(λ)  — Urbach tail')
    ax_k.set_xlabel('Wavelength (nm)'); ax_k.set_ylabel('k')
    ax_k.set_ylim(bottom=0); ax_k.legend(); ax_k.grid(True, alpha=0.2)

    plt.tight_layout()
    plt.savefig(out_png, dpi=150)
    print(f"Figure → {out_png}")

    if out_csv:
        with open(out_csv, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(['wavelength_nm', 'no', 'ne', 'k'])
            for wl_i, noi, nei, ki in zip(wl, no, ne, k):
                w.writerow([f'{wl_i:.3f}', f'{noi:.6f}', f'{nei:.6f}', f'{ki:.6f}'])
        print(f"no,ne,k table → {out_csv}")


# ══════════════════════════════════════════════════════════════
# 11. General Oscillator (Gen-Osc) model
#     Supports: Tauc-Lorentz, Gaussian oscillators
#     KK consistent by construction → physical n, k at all wavelengths
# ══════════════════════════════════════════════════════════════

# ── Oscillator primitives ──────────────────────────────────
def _tauc_lorentz_eps2(E, A, E0, C, Eg):
    """Tauc-Lorentz ε₂ (Jellison & Modine 1996).
    ε₂ = A·E0·C·(E-Eg)²/[(E²-E0²)²+C²E²] / E  for E > Eg, else 0.
    """
    E = np.asarray(E, dtype=float)
    denom = (E**2 - E0**2)**2 + C**2 * E**2
    eps2 = np.where(E > Eg,
                    A * E0 * C * (E - Eg)**2 / (denom * np.maximum(E, 1e-10)),
                    0.0)
    return eps2


def _gaussian_eps2(E, A, Ecen, Br):
    """Gaussian oscillator ε₂ (amp=A, center=Ecen eV, broadening=Br eV).
    ε₂ = A·[exp(-((E-Ecen)/Br)²) - exp(-((E+Ecen)/Br)²)]
    Antisymmetric extension ensures KK (imaginary part of Lorentzian-like).
    """
    E = np.asarray(E, dtype=float)
    return A * (np.exp(-((E - Ecen) / Br)**2) - np.exp(-((E + Ecen) / Br)**2))


# Precomputed KK operators on fixed energy grid (built once at import).
# Singularity-subtracted principal value:
#   eps1(Ei)-1 = (2/π)[ dE·Σ_j≠i (Ej·eps2_j − Ei·eps2_i)/(Ej²−Ei²)
#                       + dE·(eps2_i + Ei·eps2'_i)/(2Ei)          (j=i limit)
#                       + Ei·eps2_i · ∫_grid dE'/(E'²−Ei²) ]      (analytic)
# Validated vs analytic Lorentz: max|err| 2e-3 (vs 8e-2 for naive point-skip).
_E_KK  = np.linspace(0.5, 9.0, 1500)   # 1500-pt grid: 0.5→9 eV
_dE_KK = _E_KK[1] - _E_KK[0]

def _build_kk_ops(E):
    Ej = E[np.newaxis, :]; Ei = E[:, np.newaxis]
    with np.errstate(divide='ignore', invalid='ignore'):
        M = Ej / (Ej**2 - Ei**2)
        S = 1.0 / (Ej**2 - Ei**2)
    np.fill_diagonal(M, 0.0); np.fill_diagonal(S, 0.0)
    Ssum = S.sum(axis=1)
    a, b = E[0], E[-1]
    with np.errstate(divide='ignore'):
        Iana = (1.0/(2*E)) * (np.log(np.abs((b-E)/(b+E))) - np.log(np.abs((a-E)/(a+E))))
    # pole exactly on the integration boundary → log(0); eps2≈0 there, so drop the term
    Iana[~np.isfinite(Iana)] = 0.0
    return M, Ssum, Iana

_KK_M, _KK_Ssum, _KK_Iana = _build_kk_ops(_E_KK)

def _kk_eps1m1(eps2):
    """eps1 - 1 from eps2 on _E_KK via singularity-subtracted KK."""
    E = _E_KK
    L = (eps2 + E * np.gradient(eps2, E)) / (2*E)
    return (2/np.pi) * (_dE_KK*(_KK_M @ eps2) - _dE_KK*E*eps2*_KK_Ssum
                        + _dE_KK*L + E*eps2*_KK_Iana)


def genosc_nk(wl_nm, osc_list, eps_inf=1.0):
    """Compute n(λ), k(λ) from a list of Gen-Osc oscillators.

    osc_list: list of dicts
        TL:       {'type':'TL',       'A':, 'E0':, 'C':, 'Eg':}
        Gaussian: {'type':'Gaussian', 'A':, 'Ecen':, 'Br':}

    KK via precomputed matrix → fast enough for optimization.
    Returns: n (real), k (real), shape (len(wl_nm),)
    """
    eps2_grid = np.zeros(len(_E_KK))
    for osc in osc_list:
        if osc['type'] == 'TL':
            eps2_grid += _tauc_lorentz_eps2(_E_KK, osc['A'], osc['E0'], osc['C'], osc['Eg'])
        elif osc['type'] == 'Gaussian':
            eps2_grid += _gaussian_eps2(_E_KK, osc['A'], osc['Ecen'], osc['Br'])
    eps2_grid = np.maximum(eps2_grid, 0.0)

    eps1_grid = 1.0 + _kk_eps1m1(eps2_grid) + (eps_inf - 1.0)

    E_wl = _HC / wl_nm
    eps1_wl = PchipInterpolator(_E_KK, eps1_grid)(E_wl)
    eps2_wl = PchipInterpolator(_E_KK, eps2_grid)(E_wl)

    eps_c = eps1_wl + 1j * np.maximum(eps2_wl, 0.0)
    nc    = np.sqrt(eps_c.astype(complex))
    n = np.where(nc.real > 0, nc.real, 1.0)
    k = np.where(nc.imag > 0, nc.imag, 0.0)
    return n, k


# ── Isotropic Gen-Osc fit ──────────────────────────────────
def fit_genosc(filepath, angles_deg,
               wl_min=380, wl_max=900,
               d_film_range=(15, 60),
               d_sio2_fixed=1.5,
               d_film_init=30.0,
               # TL oscillator seed: (A, E0_eV, C_eV, Eg_eV)
               osc_seeds=None,
               eps_inf_init=1.5,
               n_lo=1.0, n_hi=3.0,
               label='Film'):
    """Isotropic Gen-Osc (Tauc-Lorentz + optional Gaussian) fit.
    osc_seeds: list of dicts, same format as genosc_nk osc_list.
    """
    if osc_seeds is None:
        osc_seeds = [{'type': 'TL', 'A': 50.0, 'E0': 3.8, 'C': 0.8, 'Eg': 3.3}]

    wl, psi_m, del_m = _load(filepath, angles_deg, wl_min, wl_max)
    N_ang = len(angles_deg)

    def unpack(x):
        d_film   = x[0]
        eps_inf  = max(x[1], 1.0)
        idx = 2
        oscs = []
        for seed in osc_seeds:
            if seed['type'] == 'TL':
                A, E0, C, Eg = x[idx], x[idx+1], x[idx+2], x[idx+3]
                A  = max(A, 0.0); C = max(C, 0.01); Eg = max(Eg, 0.1)
                E0 = max(E0, Eg + 0.05)
                oscs.append({'type': 'TL', 'A': A, 'E0': E0, 'C': C, 'Eg': Eg})
                idx += 4
            elif seed['type'] == 'Gaussian':
                A, Ecen, Br = x[idx], x[idx+1], x[idx+2]
                A = max(A, 0.0); Br = max(Br, 0.01)
                oscs.append({'type': 'Gaussian', 'A': A, 'Ecen': Ecen, 'Br': Br})
                idx += 3
        return d_film, eps_inf, oscs

    def build_x0(d_film):
        x0 = [d_film, eps_inf_init]
        for seed in osc_seeds:
            if seed['type'] == 'TL':
                x0 += [seed['A'], seed['E0'], seed['C'], seed['Eg']]
            elif seed['type'] == 'Gaussian':
                x0 += [seed['A'], seed['Ecen'], seed['Br']]
        return np.array(x0, dtype=float)

    def residuals(x):
        d_film, eps_inf, oscs = unpack(x)
        if not (d_film_range[0] <= d_film <= d_film_range[1]):
            return 1e6
        try:
            n_f, k_f = genosc_nk(wl, oscs, eps_inf=eps_inf)
        except Exception:
            return 1e6
        if np.any(n_f < n_lo) or np.any(n_f > n_hi):
            return 1e6
        N_film = (n_f + 1j * k_f).astype(complex)
        psi_c, del_c = calc_psi_delta(wl, d_sio2_fixed, d_film, N_film, angles_deg)
        dd = (del_c - del_m + 180) % 360 - 180
        r = (np.sum((psi_c - psi_m)**2) + np.sum(dd**2)) / (2 * psi_m.size)
        return r

    best_res, best_x = np.inf, None
    for d_try in np.linspace(d_film_range[0], d_film_range[1], 8):
        x0 = build_x0(d_try)
        try:
            r = minimize(residuals, x0, method='Nelder-Mead',
                         options={'maxiter': 20000, 'xatol': 1e-6, 'fatol': 1e-8,
                                  'adaptive': True})
            if r.fun < best_res:
                best_res = r.fun; best_x = r.x
        except Exception:
            pass

    res = minimize(residuals, best_x, method='Nelder-Mead',
                   options={'maxiter': 80000, 'xatol': 1e-7, 'fatol': 1e-9, 'adaptive': True})
    d_film, eps_inf, oscs = unpack(res.x)

    n_f, k_f = genosc_nk(wl, oscs, eps_inf=eps_inf)
    N_film = (n_f + 1j * k_f).astype(complex)
    psi_c, del_c = calc_psi_delta(wl, d_sio2_fixed, d_film, N_film, angles_deg)
    _dd = (del_c - del_m + 180) % 360 - 180
    rmse_psi = np.sqrt(np.mean((psi_c - psi_m)**2))
    rmse_del = np.sqrt(np.mean(_dd**2))
    rms_tot  = np.sqrt(np.mean((psi_c - psi_m)**2 + _dd**2))

    print(f"\n{'─'*55}")
    print(f"Model    : Gen-Osc (isotropic)")
    print(f"SiO2     : {d_sio2_fixed:.3f} nm (FIXED)")
    print(f"{label:<8} : {d_film:.3f} nm")
    print(f"ε∞       : {eps_inf:.4f}")
    for i, osc in enumerate(oscs):
        if osc['type'] == 'TL':
            print(f"TL osc{i+1} : A={osc['A']:.3f}  E0={osc['E0']:.4f} eV  "
                  f"C={osc['C']:.4f} eV  Eg={osc['Eg']:.4f} eV")
        elif osc['type'] == 'Gaussian':
            print(f"Gauss{i+1}  : A={osc['A']:.3f}  Ecen={osc['Ecen']:.4f} eV  "
                  f"Br={osc['Br']:.4f} eV")
    wl_rep = [400, 500, 600, 633, 700, 900]
    n_rep  = np.interp(wl_rep, wl, n_f)
    k_rep  = np.interp(wl_rep, wl, k_f)
    print("n: " + "  ".join(f"{n:.4f}@{lam}" for lam, n in zip(wl_rep, n_rep)))
    print("k: " + "  ".join(f"{k:.5f}@{lam}" for lam, k in zip(wl_rep, k_rep)))
    print(f"RMSE Ψ = {rmse_psi:.3f}°   RMSE Δ = {rmse_del:.3f}°   √MSE = {rms_tot:.4f}°")
    print(f"{'─'*55}")

    return {'wl': wl, 'n': n_f, 'k': k_f, 'N_film': N_film,
            'd_film': d_film, 'd_sio2': d_sio2_fixed,
            'psi_c': psi_c, 'del_c': del_c,
            'psi_m': psi_m, 'del_m': del_m,
            'angles': angles_deg, 'label': label,
            'rmse_psi': rmse_psi, 'rmse_del': rmse_del,
            'model': 'GenOsc', 'osc_params': oscs, 'eps_inf': eps_inf}


# ── Anisotropic Gen-Osc fit ──────────────────────────────────
def fit_genosc_aniso(filepath, angles_deg,
                     wl_min=380, wl_max=900,
                     d_film_range=(15, 60),
                     d_sio2_fixed=1.5,
                     d_film_init=30.0,
                     # separate oscillator seeds for ordinary (o) and extraordinary (e)
                     osc_seeds_o=None,
                     osc_seeds_e=None,
                     eps_inf_o_init=1.8,
                     eps_inf_e_init=1.5,
                     n_lo=1.0, n_hi=3.0,
                     label='Film'):
    """Uniaxial anisotropic Gen-Osc: separate TL/Gaussian for no and ne.
    Optical axis perpendicular to substrate (no=in-plane, ne=out-of-plane).
    """
    if osc_seeds_o is None:
        osc_seeds_o = [{'type': 'TL', 'A': 80.0, 'E0': 3.8, 'C': 0.9, 'Eg': 3.3}]
    if osc_seeds_e is None:
        osc_seeds_e = [{'type': 'TL', 'A': 40.0, 'E0': 3.7, 'C': 0.8, 'Eg': 3.3}]

    wl, psi_m, del_m = _load(filepath, angles_deg, wl_min, wl_max)

    def _n_osc(n_params_per_osc):
        return sum(4 if s['type'] == 'TL' else 3 for s in n_params_per_osc)

    no_len = _n_osc(osc_seeds_o)
    ne_len = _n_osc(osc_seeds_e)

    def unpack(x):
        d_film    = x[0]
        eps_inf_o = max(x[1], 1.0)
        eps_inf_e = max(x[2], 1.0)
        def _parse(seeds, vec, start):
            oscs = []; idx = start
            for seed in seeds:
                if seed['type'] == 'TL':
                    A, E0, C, Eg = vec[idx], vec[idx+1], vec[idx+2], vec[idx+3]
                    A=max(A,0.); C=max(C,0.01); Eg=max(Eg,0.5)
                    E0=max(E0, Eg+0.05)
                    oscs.append({'type':'TL','A':A,'E0':E0,'C':C,'Eg':Eg}); idx+=4
                elif seed['type'] == 'Gaussian':
                    A, Ecen, Br = vec[idx], vec[idx+1], vec[idx+2]
                    A=max(A,0.); Br=max(Br,0.01)
                    oscs.append({'type':'Gaussian','A':A,'Ecen':Ecen,'Br':Br}); idx+=3
            return oscs, idx
        oscs_o, end_o = _parse(osc_seeds_o, x, 3)
        oscs_e, _     = _parse(osc_seeds_e, x, end_o)
        return d_film, eps_inf_o, eps_inf_e, oscs_o, oscs_e

    def build_x0(d_film):
        x0 = [d_film, eps_inf_o_init, eps_inf_e_init]
        for seed in osc_seeds_o:
            if seed['type'] == 'TL':   x0 += [seed['A'], seed['E0'], seed['C'], seed['Eg']]
            elif seed['type'] == 'Gaussian': x0 += [seed['A'], seed['Ecen'], seed['Br']]
        for seed in osc_seeds_e:
            if seed['type'] == 'TL':   x0 += [seed['A'], seed['E0'], seed['C'], seed['Eg']]
            elif seed['type'] == 'Gaussian': x0 += [seed['A'], seed['Ecen'], seed['Br']]
        return np.array(x0, dtype=float)

    def residuals(x):
        d_film, eps_inf_o, eps_inf_e, oscs_o, oscs_e = unpack(x)
        if not (d_film_range[0] <= d_film <= d_film_range[1]):
            return 1e6
        try:
            no_f, ko = genosc_nk(wl, oscs_o, eps_inf=eps_inf_o)
            ne_f, ke = genosc_nk(wl, oscs_e, eps_inf=eps_inf_e)
        except Exception:
            return 1e6
        if np.any(no_f < n_lo) or np.any(no_f > n_hi): return 1e6
        if np.any(ne_f < n_lo) or np.any(ne_f > n_hi): return 1e6
        # shared k: average ko and ke (both should be ~same absorption)
        k_f  = 0.5 * (ko + ke)
        no_c = (no_f + 1j * k_f).astype(complex)
        ne_c = (ne_f + 1j * k_f).astype(complex)
        psi_c, del_c = calc_psi_delta_aniso(wl, d_sio2_fixed, d_film,
                                             no_c, ne_c, angles_deg)
        dd = (del_c - del_m + 180) % 360 - 180
        return np.mean((psi_c - psi_m)**2 + dd**2)

    best_res, best_x = np.inf, None
    for d_try in np.linspace(d_film_range[0], d_film_range[1], 8):
        x0 = build_x0(d_try)
        try:
            r = minimize(residuals, x0, method='Nelder-Mead',
                         options={'maxiter': 30000, 'xatol': 1e-6, 'fatol': 1e-8,
                                  'adaptive': True})
            if r.fun < best_res:
                best_res = r.fun; best_x = r.x
        except Exception:
            pass

    res = minimize(residuals, best_x, method='Nelder-Mead',
                   options={'maxiter': 100000, 'xatol': 1e-7, 'fatol': 1e-9, 'adaptive': True})
    d_film, eps_inf_o, eps_inf_e, oscs_o, oscs_e = unpack(res.x)

    no_f, ko = genosc_nk(wl, oscs_o, eps_inf=eps_inf_o)
    ne_f, ke = genosc_nk(wl, oscs_e, eps_inf=eps_inf_e)
    k_f = 0.5 * (ko + ke)
    no_c = (no_f + 1j * k_f).astype(complex)
    ne_c = (ne_f + 1j * k_f).astype(complex)
    psi_c, del_c = calc_psi_delta_aniso(wl, d_sio2_fixed, d_film, no_c, ne_c, angles_deg)

    _dd = (del_c - del_m + 180) % 360 - 180
    rmse_psi = np.sqrt(np.mean((psi_c - psi_m)**2))
    rmse_del = np.sqrt(np.mean(_dd**2))
    rms_tot  = np.sqrt(np.mean((psi_c - psi_m)**2 + _dd**2))

    print(f"\n{'─'*60}")
    print(f"Model    : Gen-Osc Uniaxial  no/ne")
    print(f"SiO2     : {d_sio2_fixed:.3f} nm (FIXED)")
    print(f"{label:<8} : {d_film:.3f} nm")
    print(f"ε∞_o={eps_inf_o:.4f}   ε∞_e={eps_inf_e:.4f}")
    print("── Ordinary (no, in-plane) ──")
    for i, osc in enumerate(oscs_o):
        if osc['type'] == 'TL':
            print(f"  TL{i+1}: A={osc['A']:.3f}  E0={osc['E0']:.4f} eV  "
                  f"C={osc['C']:.4f} eV  Eg={osc['Eg']:.4f} eV")
        else:
            print(f"  G{i+1}: A={osc['A']:.3f}  Ecen={osc['Ecen']:.4f} eV  Br={osc['Br']:.4f} eV")
    print("── Extraordinary (ne, out-of-plane) ──")
    for i, osc in enumerate(oscs_e):
        if osc['type'] == 'TL':
            print(f"  TL{i+1}: A={osc['A']:.3f}  E0={osc['E0']:.4f} eV  "
                  f"C={osc['C']:.4f} eV  Eg={osc['Eg']:.4f} eV")
        else:
            print(f"  G{i+1}: A={osc['A']:.3f}  Ecen={osc['Ecen']:.4f} eV  Br={osc['Br']:.4f} eV")
    wl_rep = [400, 500, 600, 633, 700, 900]
    no_r = np.interp(wl_rep, wl, no_f)
    ne_r = np.interp(wl_rep, wl, ne_f)
    k_r  = np.interp(wl_rep, wl, k_f)
    print("no: " + "  ".join(f"{n:.4f}@{l}" for l,n in zip(wl_rep, no_r)))
    print("ne: " + "  ".join(f"{n:.4f}@{l}" for l,n in zip(wl_rep, ne_r)))
    print("k:  " + "  ".join(f"{k:.5f}@{l}" for l,k in zip(wl_rep, k_r)))
    print(f"Δn(633nm) = no-ne = {np.interp(633, wl, no_f-ne_f):+.4f}")
    print(f"RMSE Ψ={rmse_psi:.3f}°  RMSE Δ={rmse_del:.3f}°  √MSE={rms_tot:.4f}°")
    print(f"{'─'*60}")

    return {'wl': wl, 'no': no_f, 'ne': ne_f, 'k': k_f,
            'd_film': d_film, 'd_sio2': d_sio2_fixed,
            'psi_c': psi_c, 'del_c': del_c,
            'psi_m': psi_m, 'del_m': del_m,
            'angles': angles_deg, 'label': label,
            'rmse_psi': rmse_psi, 'rmse_del': rmse_del,
            'model': 'GenOsc-Aniso'}


# ── Gen-Osc Presets ──────────────────────────────────────────
GENOSC_PRESETS = {
    'HATCN': dict(
        angles_deg=[65.0, 70.0, 75.0],
        wl_min=380, wl_max=900,
        d_film_range=(20, 45),
        d_sio2_fixed=1.5,
        d_film_init=28.0,
        # Ordinary: 1 main TL (HATCN pi→pi* at ~3.6 eV) + 1 Gaussian UV shoulder
        osc_seeds_o=[
            {'type': 'TL',      'A': 90.0, 'E0': 3.70, 'C': 0.80, 'Eg': 3.30},
            {'type': 'Gaussian','A': 1.2,  'Ecen': 4.50, 'Br': 0.50},
        ],
        # Extraordinary: weaker TL (out-of-plane weaker pi stacking contribution)
        osc_seeds_e=[
            {'type': 'TL',      'A': 50.0, 'E0': 3.65, 'C': 0.75, 'Eg': 3.30},
            {'type': 'Gaussian','A': 0.6,  'Ecen': 4.50, 'Br': 0.50},
        ],
        eps_inf_o_init=2.0,
        eps_inf_e_init=1.6,
        n_lo=1.2, n_hi=2.6,
        label='HATCN',
    ),
    'GraHIL': dict(
        angles_deg=[65.0, 70.0, 75.0],
        wl_min=400, wl_max=1000,        # drop messy deep-UV PSS band
        d_film_range=(90, 115),         # from transparent grid-scan: d≈102 nm
        d_sio2_fixed=2.0,               # Si native oxide ~2 nm
        d_film_init=102.0,
        # PEDOT:PSS-based HIL: transparent in visible (grid scan: n≈1.43, k≈0).
        # Single UV Tauc-Lorentz pole above the measured range → normal dispersion,
        # KK-consistent, n≈1.43; weak Gaussian allows any residual blue absorption.
        osc_seeds_o=[
            {'type': 'TL',      'A': 30.0, 'E0': 5.50, 'C': 1.20, 'Eg': 3.80},
            {'type': 'Gaussian','A': 0.05, 'Ecen': 3.00, 'Br': 0.60},
        ],
        # Extraordinary (out-of-plane): PEDOT:PSS literature ne ~7-10% higher than no
        osc_seeds_e=[
            {'type': 'TL',      'A': 34.0, 'E0': 5.50, 'C': 1.20, 'Eg': 3.80},
            {'type': 'Gaussian','A': 0.05, 'Ecen': 3.00, 'Br': 0.60},
        ],
        eps_inf_o_init=1.40,
        eps_inf_e_init=1.45,
        n_lo=1.30, n_hi=1.75,
        label='GraHIL',
    ),
}


# ══════════════════════════════════════════════════════════════
# 12. Perovskite-on-GraHIL multilayer (Gen-Osc + exciton, EMA roughness/damage)
#     Stack: Air / rough-EMA / Perovskite(GenOsc) / damage-EMA / GraHIL(fixed) / SiO2 / Si
# ══════════════════════════════════════════════════════════════

def load_grahil_nk(csv_path):
    """Load GraHIL n,k dispersion from a prior fit CSV → complex N(wl) function.
    Flat edge-extrapolation outside the fitted range (GraHIL is transparent,
    low-dispersion, so this is safe and its buried value has low leverage)."""
    wls, ns, ks = [], [], []
    with open(csv_path) as f:
        r = csv.reader(f); next(r)
        for row in r:
            wls.append(float(row[0])); ns.append(float(row[1])); ks.append(float(row[2]))
    wls = np.array(wls); ns = np.array(ns); ks = np.array(ks)
    fn = interp1d(wls, ns, bounds_error=False, fill_value=(ns[0], ns[-1]))
    fk = interp1d(wls, ks, bounds_error=False, fill_value=(ks[0], ks[-1]))
    def N_grahil(wl):
        return (fn(wl) + 1j * np.maximum(fk(wl), 0.0)).astype(complex)
    return N_grahil


def load_depol(filepath, angles_deg):
    """Load % Depolarization xlsx (header rows 3-4, data row 5+; colB=wl, then
    one depol column per angle). Returns wl_d (Nd,), dep (Nd, n_ang)."""
    import openpyxl
    n_ang = len(angles_deg)
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    wl_d, rows = [], []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[1] is None or not isinstance(row[1], (int, float)):
            continue
        vals = [row[2 + j] for j in range(n_ang)]
        if any(v is None for v in vals):
            continue
        wl_d.append(float(row[1])); rows.append([float(v) for v in vals])
    wb.close()
    return np.array(wl_d), np.array(rows)


def depol_weights(wl, angles_deg, depol_file, depol_cut=3.0, depol_soft=1.5, floor=0.02):
    """Per-(wl,angle) fit weights from a depolarization spectrum.
    High |depol| flags unreliable points (e.g. Si-backside NIR) → weight≈floor;
    otherwise a smooth 1/(1+(depol/depol_soft)²) taper. Returns (Nwl, n_ang)."""
    wl_d, dep = load_depol(depol_file, angles_deg)
    W = np.ones((len(wl), len(angles_deg)))
    for j in range(len(angles_deg)):
        dj = np.interp(wl, wl_d, dep[:, j])
        soft = 1.0 / (1.0 + (dj / depol_soft)**2)
        W[:, j] = np.where(np.abs(dj) > depol_cut, floor, np.maximum(soft, floor))
    return W


def calc_psi_delta_perov(wl, N_perov, N_grahil, d_rough, d_perov,
                         d_damage, d_grahil, d_sio2, angles,
                         f_rough=0.5, f_damage=0.5):
    """Ψ/Δ for the perovskite/GraHIL stack.
    rough-EMA = f_rough perovskite + void ; damage-EMA = f_damage GraHIL + perovskite.
    Layers with ~zero thickness are dropped."""
    n_air = np.ones(len(wl), dtype=complex)
    N_gr  = N_grahil(wl)
    layers = [n_air]; d_list = []
    if d_rough > 1e-6:
        layers.append(bruggeman_ema(N_perov, n_air, f_rough)); d_list.append(d_rough)
    layers.append(N_perov); d_list.append(d_perov)
    if d_damage > 1e-6:
        layers.append(bruggeman_ema(N_gr, N_perov, f_damage)); d_list.append(d_damage)
    layers.append(N_gr);          d_list.append(d_grahil)
    layers.append(n_SiO2(wl));    d_list.append(d_sio2)
    layers.append(n_Si(wl))
    psi_out, del_out = [], []
    for ang in angles:
        rp, rs = _tmm(wl, layers, d_list, ang)
        rho = rp / rs
        psi_out.append(np.degrees(np.arctan(np.abs(rho))))
        del_out.append(np.degrees(np.angle(rho)))
    return np.array(psi_out).T, np.array(del_out).T


def _clip(v, lo, hi):
    return min(max(v, lo), hi)

def _perov_oscs(eps_inf, tl, exc, uv):
    """Build a physically-bounded oscillator list for a bromide perovskite:
    band-edge TL (Eg~2.3eV), phonon-broadened Gaussian exciton, and a UV-TL band."""
    Eg  = _clip(tl[3], 1.9, 2.4)
    uEg = _clip(uv[3], 2.3, 3.8)
    return [
        {'type': 'TL',       'A': _clip(tl[0], 0.0, 300.0),
         'E0': _clip(tl[1], Eg+0.05, 3.2), 'C': _clip(tl[2], 0.05, 1.5), 'Eg': Eg},
        {'type': 'Gaussian', 'A': _clip(exc[0], 0.0, 5.0),
         'Ecen': _clip(exc[1], 2.20, 2.45), 'Br': _clip(exc[2], 0.03, 0.08)},
        {'type': 'TL',       'A': _clip(uv[0], 0.0, 400.0),
         'E0': _clip(uv[1], uEg+0.05, 5.5), 'C': _clip(uv[2], 0.2, 2.5), 'Eg': uEg},
    ]


def fit_perovskite(filepath, angles_deg,
                   wl_min=380, wl_max=1688,
                   wl_transp=650.0,          # Stage-1 transparent-region lower bound
                   d_sio2_fixed=2.0,
                   grahil_csv=None,
                   d_perov_meas=270.0,       # SEM thickness prior (measured, soft)
                   d_rough_seed=28.0, d_rough_range=(3.0, 55.0),
                   d_intermix_seed=30.0, d_intermix_range=(0.0, 120.0),
                   f_intermix=0.5,           # GraHIL fraction in the damage intermix EMA
                   eps_inf_init=2.0,
                   tl_seed=(30.0, 2.45, 0.30, 2.28),
                   exc_seed=(0.20, 2.33, 0.05),
                   uv_seed=(75.0, 3.75, 1.20, 2.50),
                   reg_thick=0.02, reg_ktransp=50.0,
                   depol_file=None, depol_cut=3.0, depol_soft=1.5,
                   nir_weight=0.3, nir_full=750.0, nir_low=1050.0,
                   label='Perovskite'):
    """Two-stage best-effort fit of perovskite on a damaged GraHIL.
    The buried GraHIL is chemically consumed by the DMSO precursor, so it is
    represented by a single intermix EMA (GraHIL+perovskite) directly on SiO2/Si.
    If depol_file is given, per-point weights come from the measured
    depolarization (high |depol| → unreliable, e.g. Si-backside NIR>1100nm →
    down-weighted); otherwise a geometric NIR taper is used.
    Stack: Air / rough-EMA / Perov(GenOsc) / intermix-EMA / SiO2 / Si."""
    if grahil_csv is None:
        grahil_csv = str(OUT_DIR / 'GraHIL_genosc_nk.csv')
    N_grahil = load_grahil_nk(grahil_csv)

    wl, psi_m, del_m = _load(filepath, angles_deg, wl_min, wl_max)

    # per-(wl,angle) weights: prefer measured depolarization; else geometric NIR taper
    if depol_file is not None:
        W = depol_weights(wl, angles_deg, depol_file, depol_cut=depol_cut, depol_soft=depol_soft)
        print(f"[weights] depol-based: mean w={W.mean():.3f}, "
              f"{100*np.mean(W < 0.1):.1f}% of points strongly down-weighted (|depol|>{depol_cut}%)")
    else:
        ramp = (wl - nir_full) / (nir_low - nir_full)
        w = np.where(wl <= nir_full, 1.0,
            np.where(wl >= nir_low, nir_weight, 1.0 - (1.0 - nir_weight)*np.clip(ramp, 0, 1)))
        W = w[:, None] * np.ones((1, len(angles_deg)))

    def _model(N_p, d_rough, d_perov, d_intermix, wl_arr=None, mask=None):
        wla = wl if wl_arr is None else wl_arr
        return calc_psi_delta_perov(wla, N_p, N_grahil, d_rough, d_perov,
                                    d_intermix, 0.0, d_sio2_fixed, angles_deg,
                                    f_damage=f_intermix)

    # ---------- Stage 1: transparent Cauchy, k=0 (weighted) ----------
    tmask = wl >= wl_transp
    wlt, psit, delt, Wt = wl[tmask], psi_m[tmask], del_m[tmask], W[tmask]

    def s1_unpack(x):
        d_rough    = min(max(x[0], d_rough_range[0]), d_rough_range[1])
        d_perov    = min(max(x[1], 180.0), 340.0)
        d_intermix = min(max(x[2], d_intermix_range[0]), d_intermix_range[1])
        cA = min(max(x[3], 1.6), 2.7)
        cB = min(max(x[4], 0.0), 80000.0)
        return d_rough, d_perov, d_intermix, cA, cB

    def s1_res(x):
        d_rough, d_perov, d_intermix, cA, cB = s1_unpack(x)
        N_p = ((cA + cB / wlt**2) + 0j).astype(complex)
        psi_c, del_c = calc_psi_delta_perov(wlt, N_p, N_grahil, d_rough, d_perov,
                                            d_intermix, 0.0, d_sio2_fixed, angles_deg,
                                            f_damage=f_intermix)
        dd = (del_c - delt + 180) % 360 - 180
        mse = (np.sum(Wt*(psi_c - psit)**2) + np.sum(Wt*dd**2)) / (2 * np.sum(Wt))
        prior = reg_thick * (d_perov + 0.5*d_rough - d_perov_meas)**2
        return mse + prior

    best1, bx1 = np.inf, None
    for dp in np.linspace(235, 285, 6):
        for dr in (10, 28, 45):
            for di in (5, 40):
                r = minimize(s1_res, [dr, dp, di, 2.15, 15000.0], method='Nelder-Mead',
                             options={'maxiter': 8000, 'xatol': 1e-5, 'fatol': 1e-8, 'adaptive': True})
                if r.fun < best1:
                    best1, bx1 = r.fun, r.x
    d_rough1, d_perov1, d_int1, cA1, cB1 = s1_unpack(bx1)
    print(f"[Stage 1] transparent(weighted): d_perov={d_perov1:.1f} d_rough={d_rough1:.1f} "
          f"intermix={d_int1:.1f} nm  n(700)={cA1+cB1/700**2:.3f} n(1000)={cA1+cB1/1000**2:.3f} "
          f"(res={best1:.4f})")

    # ---------- Stage 2: full-range Gen-Osc (weighted) ----------
    dr_lo, dr_hi = d_rough_range
    dp_lo, dp_hi = d_perov1*0.90, d_perov1*1.10
    di_lo, di_hi = d_intermix_range

    def s2_unpack(x):
        d_rough    = min(max(x[0], dr_lo), dr_hi)
        d_perov    = min(max(x[1], dp_lo), dp_hi)
        d_intermix = min(max(x[2], di_lo), di_hi)
        eps_inf = max(x[3], 1.0)
        tl  = (x[4], x[5], x[6], x[7]); exc = (x[8], x[9], x[10]); uv = (x[11], x[12], x[13], x[14])
        return d_rough, d_perov, d_intermix, eps_inf, tl, exc, uv

    def s2_res(x):
        d_rough, d_perov, d_intermix, eps_inf, tl, exc, uv = s2_unpack(x)
        oscs = _perov_oscs(eps_inf, tl, exc, uv)
        try:
            n_p, k_p = genosc_nk(wl, oscs, eps_inf=eps_inf)
        except Exception:
            return 1e6
        if np.any(n_p < 1.5) or np.any(n_p > 2.75):
            return 1e6
        N_p = (n_p + 1j * k_p).astype(complex)
        psi_c, del_c = calc_psi_delta_perov(wl, N_p, N_grahil, d_rough, d_perov,
                                            d_intermix, 0.0, d_sio2_fixed, angles_deg,
                                            f_damage=f_intermix)
        dd = (del_c - del_m + 180) % 360 - 180
        mse = (np.sum(W*(psi_c - psi_m)**2) + np.sum(W*dd**2)) / (2 * np.sum(W))
        prior = reg_thick * (d_perov + 0.5*d_rough - d_perov_meas)**2
        ktransp = reg_ktransp * np.mean((k_p[wl > 640.0])**2)   # enforce sub-gap k≈0
        return mse + prior + ktransp

    x0 = [d_rough1, d_perov1, d_int1, eps_inf_init,
          tl_seed[0], tl_seed[1], tl_seed[2], tl_seed[3],
          exc_seed[0], exc_seed[1], exc_seed[2],
          uv_seed[0], uv_seed[1], uv_seed[2], uv_seed[3]]
    res = minimize(s2_res, x0, method='Nelder-Mead',
                   options={'maxiter': 80000, 'xatol': 1e-6, 'fatol': 1e-9, 'adaptive': True})
    res = minimize(s2_res, res.x, method='Nelder-Mead',
                   options={'maxiter': 80000, 'xatol': 1e-7, 'fatol': 1e-10, 'adaptive': True})
    d_rough, d_perov, d_intermix, eps_inf, tl, exc, uv = s2_unpack(res.x)
    oscs = _perov_oscs(eps_inf, tl, exc, uv)
    n_p, k_p = genosc_nk(wl, oscs, eps_inf=eps_inf)
    N_p = (n_p + 1j * k_p).astype(complex)
    psi_c, del_c = calc_psi_delta_perov(wl, N_p, N_grahil, d_rough, d_perov,
                                        d_intermix, 0.0, d_sio2_fixed, angles_deg,
                                        f_damage=f_intermix)
    _dd = (del_c - del_m + 180) % 360 - 180
    # unweighted RMSE (all λ) and weighted-region RMSE (λ<nir_low, the reliable band)
    rmse_psi = np.sqrt(np.mean((psi_c - psi_m)**2))
    rmse_del = np.sqrt(np.mean(_dd**2))
    rms_tot  = np.sqrt(np.mean((psi_c - psi_m)**2 + _dd**2))
    rel = wl < nir_low
    rms_rel = np.sqrt(np.mean((psi_c[rel]-psi_m[rel])**2 + _dd[rel]**2))

    wl_rep = [420, 460, 500, 535, 550, 600, 700, 900]
    n_rep = np.interp(wl_rep, wl, n_p); k_rep = np.interp(wl_rep, wl, k_p)
    print(f"\n{'─'*62}")
    print(f"Model    : Perovskite Gen-Osc (TL + Gaussian exciton + UV-TL)")
    print(f"Stack    : Air / rough-EMA / Perov / intermix-EMA(GraHIL+Perov) / SiO2 / Si")
    print(f"SiO2     : {d_sio2_fixed:.2f} nm (fixed)   GraHIL bulk: consumed (0 nm)")
    print(f"rough-EMA: {d_rough:.2f} nm    Perov bulk: {d_perov:.2f} nm    "
          f"intermix-EMA: {d_intermix:.2f} nm (f_GraHIL={f_intermix:.2f})")
    print(f"total(bulk+½rough) = {d_perov + 0.5*d_rough:.1f} nm  (SEM prior {d_perov_meas:.0f})")
    print(f"ε∞={eps_inf:.3f}")
    print(f"TL   : A={tl[0]:.2f} E0={tl[1]:.3f} C={tl[2]:.3f} Eg={tl[3]:.3f} eV "
          f"({_HC/max(tl[3],0.1):.0f} nm edge)")
    print(f"Exc  : A={exc[0]:.3f} Ecen={exc[1]:.3f} eV ({_HC/exc[1]:.0f} nm) Br={min(max(exc[2],0.03),0.08):.3f} eV")
    print(f"UV-TL: A={uv[0]:.2f} E0={uv[1]:.3f} C={uv[2]:.3f} Eg={uv[3]:.3f} eV")
    print("n: " + " ".join(f"{n:.3f}@{l}" for l, n in zip(wl_rep, n_rep)))
    print("k: " + " ".join(f"{k:.3f}@{l}" for l, k in zip(wl_rep, k_rep)))
    print(f"RMSE(all λ) Ψ={rmse_psi:.3f}°  Δ={rmse_del:.3f}°  √MSE={rms_tot:.4f}°")
    print(f"RMSE(λ<{nir_low:.0f}nm, reliable band) √MSE={rms_rel:.4f}°")
    print(f"{'─'*62}")

    return {'wl': wl, 'n': n_p, 'k': k_p, 'N_film': N_p,
            'd_film': d_perov, 'd_sio2': d_sio2_fixed,
            'd_rough': d_rough, 'd_damage': d_intermix, 'd_grahil': 0.0,
            'psi_c': psi_c, 'del_c': del_c, 'psi_m': psi_m, 'del_m': del_m,
            'angles': angles_deg, 'label': label,
            'rmse_psi': rmse_psi, 'rmse_del': rmse_del, 'rms_reliable': rms_rel,
            'model': 'Perovskite-GenOsc', 'osc_params': oscs, 'eps_inf': eps_inf}


PEROVSKITE_PRESETS = {
    'Perov': dict(
        angles_deg=[65.0, 70.0, 75.0],
        wl_min=380, wl_max=1688,
        wl_transp=650.0,
        d_sio2_fixed=2.0,
        d_perov_meas=270.0,                          # SEM prior (soft)
        d_rough_seed=40.0, d_rough_range=(3.0, 70.0),
        d_intermix_seed=20.0, d_intermix_range=(0.0, 120.0),
        f_intermix=0.5,                              # GraHIL+perovskite damage intermix
        eps_inf_init=2.0,
        tl_seed=(30.0, 2.45, 0.30, 2.28),
        exc_seed=(0.30, 2.33, 0.05),
        uv_seed=(75.0, 3.75, 1.20, 2.50),
        reg_thick=0.08,                              # mild SEM anchor (data now clean via depol mask)
        depol_file=str(DATA_DIR / '#2_depol.xlsx'),
        depol_cut=3.0, depol_soft=1.5,
        nir_weight=0.15, nir_full=700.0, nir_low=1000.0,
        label='Perovskite',
    ),
}


def plot_results_perov(r, out_png, out_csv=None):
    """Plot perovskite stack fit: Ψ, Δ, n, k with thickness annotations."""
    wl = r['wl']; psi_m = r['psi_m']; del_m = r['del_m']
    psi_c = r['psi_c']; del_c = r['del_c']
    n = r['n']; k = r['k']; angles = r['angles']; label = r['label']
    n_ang = len(angles)
    cmap = matplotlib.colormaps['tab10']
    colors = [cmap(i/9.0) for i in range(n_ang)]
    del_c_plot, del_m_plot = _align_delta(del_c, del_m)

    fig, axes = plt.subplots(2, 2, figsize=(14, 9))
    fig.suptitle(f'{label} on GraHIL — Ellipsometry Fit (Gen-Osc + exciton)\n'
                 f'rough={r["d_rough"]:.1f}nm  bulk={r["d_film"]:.1f}nm  '
                 f'damage-EMA={r["d_damage"]:.1f}nm  GraHIL={r["d_grahil"]:.0f}nm(fix)  '
                 f'√MSE={np.sqrt(r["rmse_psi"]**2+r["rmse_del"]**2):.3f}°',
                 fontsize=9, fontweight='bold')
    ax_p, ax_d, ax_n, ax_k = axes[0,0], axes[0,1], axes[1,0], axes[1,1]

    for i, (ang, col) in enumerate(zip(angles, colors)):
        ax_p.plot(wl, psi_m[:,i], 'o', ms=1.5, color=col, alpha=0.35)
        ax_p.plot(wl, psi_c[:,i], '-', lw=1.2, color=col, label=f'{ang:.0f}°')
        ax_d.plot(wl, del_m_plot[:,i], 'o', ms=1.5, color=col, alpha=0.35)
        ax_d.plot(wl, del_c_plot[:,i], '-', lw=1.2, color=col, label=f'{ang:.0f}°')
    for ax, title, ylab in [(ax_p,'Ψ (Psi)','Ψ (deg)'), (ax_d,'Δ (Delta)','Δ (deg)')]:
        ax.set_title(title); ax.set_xlabel('Wavelength (nm)'); ax.set_ylabel(ylab)
        ax.legend(fontsize=7); ax.grid(True, alpha=0.2)

    ax_n.plot(wl, n, 'b-', lw=1.5)
    ax_n.axvline(_HC/2.30, color='gray', ls=':', lw=0.8)
    ax_n.set_title(f'{label}  n(λ)'); ax_n.set_xlabel('Wavelength (nm)'); ax_n.set_ylabel('n')
    ax_n.grid(True, alpha=0.2)
    ax_k.plot(wl, k, 'r-', lw=1.5)
    ax_k.axvline(550, color='green', ls=':', lw=0.8, label='550 nm (k→0)')
    ax_k.set_title(f'{label}  k(λ)'); ax_k.set_xlabel('Wavelength (nm)'); ax_k.set_ylabel('k')
    ax_k.set_ylim(bottom=0); ax_k.legend(fontsize=7); ax_k.grid(True, alpha=0.2)

    plt.tight_layout(); plt.savefig(out_png, dpi=150)
    print(f"Figure → {out_png}")
    if out_csv:
        with open(out_csv, 'w', newline='') as f:
            w = csv.writer(f); w.writerow(['wavelength_nm', 'n', 'k'])
            for wl_i, ni, ki in zip(wl, n, k):
                w.writerow([f'{wl_i:.3f}', f'{ni:.6f}', f'{ki:.6f}'])
        print(f"n,k table → {out_csv}")


# ══════════════════════════════════════════════════════════════
# 13. Main
# ══════════════════════════════════════════════════════════════
if __name__ == '__main__':
    import sys
    material = sys.argv[1] if len(sys.argv) > 1 else 'Bphen'
    _default_mode = {'HATCN': 'genosc_aniso', 'GraHIL': 'genosc',
                     'Perov': 'perovskite'}.get(material, 'bspline_n')
    mode     = sys.argv[2] if len(sys.argv) > 2 else _default_mode
    filepath = sys.argv[3] if len(sys.argv) > 3 else {
        'Bphen':  str(DATA_DIR / 'Bphen.txt'),
        'HATCN':  str(DATA_DIR / '#1.xlsx'),
        'GraHIL': str(DATA_DIR / '#1.xlsx'),
        'Perov':  str(DATA_DIR / '#2.xlsx'),
        'MgAg':   str(OUT_DIR / 'MgAg1.txt'),
    }.get(material, str(OUT_DIR / 'MgAg1.txt'))

    out_base = str(OUT_DIR / material)

    if mode == 'perovskite' and material in PEROVSKITE_PRESETS:
        r = fit_perovskite(filepath, **PEROVSKITE_PRESETS[material])
        plot_results_perov(r, out_png=out_base + '_perov_fit.png',
                           out_csv=out_base + '_perov_nk.csv')
    elif mode == 'genosc_aniso' and material in GENOSC_PRESETS:
        r = fit_genosc_aniso(filepath, **GENOSC_PRESETS[material])
        plot_results_aniso(r,
                           out_png=out_base + '_genosc_aniso_fit.png',
                           out_csv=out_base + '_genosc_aniso_nk.csv')
    elif mode == 'genosc' and material in GENOSC_PRESETS:
        # build isotropic preset from aniso preset
        p = {k: v for k, v in GENOSC_PRESETS[material].items()
             if k not in ('osc_seeds_o', 'osc_seeds_e', 'eps_inf_o_init', 'eps_inf_e_init')}
        p['osc_seeds'] = GENOSC_PRESETS[material]['osc_seeds_o']
        p['eps_inf_init'] = GENOSC_PRESETS[material]['eps_inf_o_init']
        r = fit_genosc(filepath, **p)
        plot_results(r, out_png=out_base + '_genosc_fit.png',
                     out_csv=out_base + '_genosc_nk.csv')
    elif mode == 'aniso' and material in ANISO_PRESETS:
        r = fit_bspline_aniso(filepath, **ANISO_PRESETS[material])
        plot_results_aniso(r,
                           out_png=out_base + '_aniso_fit.png',
                           out_csv=out_base + '_aniso_nk.csv')
    elif mode == 'bspline_n' and material in BSPLINE_N_PRESETS:
        r = fit_bspline_n(filepath, **BSPLINE_N_PRESETS[material])
        plot_results(r, out_png=out_base+'_fit.png', out_csv=out_base+'_nk.csv')
    elif mode == 'cauchy' and material in CAUCHY_PRESETS:
        r = fit_cauchy(filepath, **CAUCHY_PRESETS[material])
        plot_results(r, out_png=out_base+'_fit.png', out_csv=out_base+'_nk.csv')
    else:
        r = fit(filepath, **PRESETS[material])
        plot_results(r, out_png=out_base+'_fit.png', out_csv=out_base+'_nk.csv')
